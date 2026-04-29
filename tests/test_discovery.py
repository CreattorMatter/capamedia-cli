from __future__ import annotations

import subprocess
from pathlib import Path

from openpyxl import Workbook

from capamedia_cli.commands.discovery import (
    _copy_spec_artifacts,
    _default_edge_case_report,
    _local_spec_artifact_destination,
    _probe_spec_repo,
)
from capamedia_cli.core.discovery import (
    DISCOVERY_WORKBOOK_NAME,
    DiscoveryEntry,
    DiscoverySpecArtifact,
    DiscoverySpecProbe,
    bundled_discovery_workbook,
    classify_edge_cases,
    detect_discovery_workspace,
    find_discovery_workbook,
    load_discovery_entry,
    parse_azure_path,
    parse_azure_repo_name,
    rank_spec_candidate,
    render_discovery_markdown,
    service_suffix_key,
    spec_parent_path,
)


def test_bundled_discovery_workbook_exists() -> None:
    workbook = bundled_discovery_workbook()

    assert workbook is not None
    assert workbook.name == DISCOVERY_WORKBOOK_NAME


def _make_discovery(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Validacion de servicios"
    ws.append(
        [
            "Servicio",
            "Nuevo nombre",
            "TRIBU",
            "ACRONIMO",
            "Tecnologia",
            "Tipo",
            "Integraciones / Consume",
            "Cache Adicional al config",
            "Archivo o servicio de donde obtiene informacion para cache",
            "Interaccion con proveedores externos",
            "Metodos que expone",
            "Peso del servicio",
            "Complejidad del servicio",
            "Observacion Discovery",
            "LINK WSDL",
            "LINK CODIGO",
            "Consumen tecnologia deprecada",
            "Peso",
        ]
    )
    ws.append(
        [
            "WSClientes0028",
            "tnd-msa-sp-wsclientes0028",
            "TRIBU SEGMENTOS Y NEGOCIOS DIGITALES",
            "tnd",
            "Bus Omnicanalidad",
            "WS",
            "UMPClientes0020 -> TX067050\nServicio Configurable en WS y Ump",
            "Uso de cache",
            "configuraciones/000000001.xml",
            "Ninguno",
            "ActualizarCelularLocalizacion34\nActualizarEmailLocalizacion33",
            "13",
            "Bajo",
            "Validar las descripciones de las tx.",
            "specs",
            "code",
            "PCBG-999",
            "Alta",
        ]
    )
    ws["O2"].hyperlink = (
        "https://dev.azure.com/BancoPichinchaEC/adi-especificaciones-tecnicas/"
        "_git/adi-doc-tecspec-tribu-integracion-apis?path=/sp%20-%20Soporte/"
        "tnd-msa-sp-wsclientes0028"
    )
    ws["P2"].hyperlink = (
        "https://dev.azure.com/BancoPichinchaEC/tpl-bus-omnicanal/"
        "_git/sqb-msa-wsclientes0028"
    )
    wb.save(path)


def test_parse_azure_repo_and_path() -> None:
    url = (
        "https://dev.azure.com/BancoPichinchaEC/adi-especificaciones-tecnicas/"
        "_git/adi-doc-tecspec-tribu-integracion-apis?version=GBmaster&"
        "path=/sp%20-%20Soporte/tnd-msa-sp-wsclientes0028"
    )

    assert parse_azure_repo_name(url) == "adi-doc-tecspec-tribu-integracion-apis"
    assert parse_azure_path(url) == "/sp - Soporte/tnd-msa-sp-wsclientes0028"


def test_spec_path_fuzzy_helpers_handle_acronym_rename() -> None:
    broken_path = "/sp - Soporte/mdw-msa-sp-wsreglas0010"

    assert spec_parent_path(broken_path) == "sp - Soporte"
    assert service_suffix_key(broken_path) == "wsreglas0010"
    assert rank_spec_candidate("csg-msa-sp-wsreglas0010", "wsreglas0010", "mdw") > 0
    assert rank_spec_candidate("csg-msa-sp-wsclientes0010", "wsreglas0010", "mdw") == 0


def test_probe_spec_repo_falls_back_to_service_suffix_and_copies_artifacts(tmp_path: Path) -> None:
    source_repo = tmp_path / "source-specs"
    spec_dir = source_repo / "sp - Soporte" / "csg-msa-sp-wsreglas0010"
    spec_dir.mkdir(parents=True)
    for name in ("GenericSOAP.xsd", "WSReglas0010_InlineSchema1.xsd", "WSReglas0010.wsdl"):
        (spec_dir / name).write_text("<xml/>", encoding="utf-8")

    subprocess.run(["git", "init"], cwd=source_repo, check=True, capture_output=True, text=True)
    subprocess.run(["git", "config", "user.email", "test@example.com"], cwd=source_repo, check=True)
    subprocess.run(["git", "config", "user.name", "Test"], cwd=source_repo, check=True)
    subprocess.run(["git", "add", "."], cwd=source_repo, check=True)
    subprocess.run(["git", "commit", "-m", "specs"], cwd=source_repo, check=True, capture_output=True, text=True)

    entry = DiscoveryEntry(
        service="WSReglas0010",
        migrated_name="mdw-msa-sp-wsreglas0010",
        acronym="mdw",
        link_wsdl=source_repo.resolve().as_uri() + "?path=/sp%20-%20Soporte/mdw-msa-sp-wsreglas0010",
        spec_repo="spec",
        spec_path="/sp - Soporte/mdw-msa-sp-wsreglas0010",
    )

    probe = _probe_spec_repo(entry, tmp_path / "cache")

    assert probe.status == "ok"
    assert probe.requested_path == "/sp - Soporte/mdw-msa-sp-wsreglas0010"
    assert probe.resolved_path == "/sp - Soporte/csg-msa-sp-wsreglas0010"
    assert {artifact.path.name for artifact in probe.artifacts} == {
        "GenericSOAP.xsd",
        "WSReglas0010_InlineSchema1.xsd",
        "WSReglas0010.wsdl",
    }

    destination = tmp_path / ".capamedia" / "discovery" / "wsreglas0010" / "specs"
    copied = _copy_spec_artifacts(probe, destination)

    assert {path.name for path in copied} == {
        "GenericSOAP.xsd",
        "WSReglas0010_InlineSchema1.xsd",
        "WSReglas0010.wsdl",
    }
    assert all(destination in path.parents for path in copied)


def test_classify_edge_cases_detects_discovery_patterns() -> None:
    cases = classify_edge_cases(
        integrations="CE_EVENTOS\nUMPClientes0020 -> TX067050",
        observations="No hay las fuentes de la UMPClientes0161. Validar las descripciones de las tx.",
        cache="Persiste datos en cache Shared Row",
    )

    codes = {case.code for case in cases}
    assert "mq_or_event" in codes
    assert "same_name_bus_was_or_missing_source" in codes
    assert "tx_description_validation" in codes
    assert "cache_or_config_file" in codes


def test_load_discovery_entry_by_service_and_render_markdown(tmp_path: Path) -> None:
    xlsx = tmp_path / "discovery.xlsx"
    _make_discovery(xlsx)

    entry = load_discovery_entry(xlsx, "wsclientes0028")

    assert entry is not None
    assert entry.migrated_name == "tnd-msa-sp-wsclientes0028"
    assert entry.spec_repo == "adi-doc-tecspec-tribu-integracion-apis"
    assert entry.spec_path == "/sp - Soporte/tnd-msa-sp-wsclientes0028"
    assert entry.code_repo == "sqb-msa-wsclientes0028"
    assert "Consumen tecnologia deprecada: Alta" in entry.weight_flags
    assert {case.code for case in entry.edge_cases} >= {
        "cache_or_config_file",
        "tx_description_validation",
    }

    copied = tmp_path / ".capamedia" / "discovery" / "wsclientes0028" / "specs" / "WSClientes0028.wsdl"
    probe = DiscoverySpecProbe(
        status="ok",
        artifacts=[
            DiscoverySpecArtifact(path=tmp_path / "GenericSOAP.xsd", kind="xsd"),
            DiscoverySpecArtifact(path=tmp_path / "WSClientes0028.wsdl", kind="wsdl"),
        ],
        resolved_path="/sp - Soporte/csg-msa-sp-wsclientes0028",
        requested_path="/sp - Soporte/mdw-msa-sp-wsclientes0028",
    )
    markdown = render_discovery_markdown(entry, spec_probe=probe, copied_artifacts=[copied])
    assert "## Discovery / edge cases" in markdown
    assert "DISCOVERY_EDGE_CASES:" in markdown
    assert "tnd-msa-sp-wsclientes0028" in markdown
    assert "Spec path resuelto" in markdown
    assert "GenericSOAP.xsd" in markdown
    assert ".capamedia" in markdown
    assert "tx_description_validation" in markdown
    assert "<pendiente_validar>" in markdown


def test_load_discovery_entry_by_migrated_name(tmp_path: Path) -> None:
    xlsx = tmp_path / "discovery.xlsx"
    _make_discovery(xlsx)

    entry = load_discovery_entry(xlsx, "tnd-msa-sp-wsclientes0028")

    assert entry is not None
    assert entry.service == "WSClientes0028"


def test_find_discovery_workbook_searches_ancestors(tmp_path: Path) -> None:
    workbook = tmp_path / DISCOVERY_WORKBOOK_NAME
    _make_discovery(workbook)
    nested = tmp_path / "wsclientes0028" / "destino" / "tnd-msa-sp-wsclientes0028"
    nested.mkdir(parents=True)

    assert find_discovery_workbook(nested) == workbook


def test_detect_discovery_workspace_from_here_structure(tmp_path: Path) -> None:
    legacy = tmp_path / "legacy" / "sqb-msa-wsclientes0028"
    migrated = tmp_path / "destino" / "tnd-msa-sp-wsclientes0028"
    legacy.mkdir(parents=True)
    migrated.mkdir(parents=True)

    ctx = detect_discovery_workspace(migrated)

    assert ctx.root == tmp_path
    assert ctx.service_name == "wsclientes0028"
    assert ctx.legacy_path == legacy
    assert ctx.migrated_path == migrated


def test_discovery_here_defaults_to_local_capamedia_artifacts(tmp_path: Path) -> None:
    legacy = tmp_path / "legacy" / "sqb-msa-wsclientes0028"
    migrated = tmp_path / "destino" / "tnd-msa-sp-wsclientes0028"
    legacy.mkdir(parents=True)
    migrated.mkdir(parents=True)

    ctx = detect_discovery_workspace(migrated)
    entry = DiscoveryEntry(service="WSClientes0028")

    assert _local_spec_artifact_destination(ctx, entry) == (
        tmp_path / ".capamedia" / "discovery" / "wsclientes0028" / "specs"
    )
    assert _default_edge_case_report(ctx, entry) == (
        tmp_path / ".capamedia" / "reports" / "discovery-edge-cases-wsclientes0028.md"
    )
