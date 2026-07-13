"""Tests for the deterministic checklist rules."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from capamedia_cli.core.autofix import run_autofix_loop
from capamedia_cli.core.checklist_rules import (
    CheckContext,
    run_block_0,
    run_block_1,
    run_block_7,
    run_block_8,
    run_block_13,
    run_block_15,
    run_block_21,
    run_block_22,
)
from capamedia_cli.core.discovery import DISCOVERY_WORKBOOK_NAME


def _make_migrated(tmp_path: Path) -> Path:
    """Create a minimal migrated project layout for tests."""
    root = tmp_path / "migrated"
    (root / "src" / "main" / "java" / "com" / "pichincha" / "sp" / "domain").mkdir(parents=True)
    (root / "src" / "main" / "java" / "com" / "pichincha" / "sp" / "application").mkdir(parents=True)
    (root / "src" / "main" / "java" / "com" / "pichincha" / "sp" / "infrastructure").mkdir(parents=True)
    (root / "src" / "main" / "resources").mkdir(parents=True)
    return root


def _make_discovery_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Validacion de servicios"
    ws.append(
        [
            "Servicio",
            "Nuevo nombre",
            "TRIBU",
            "ACRONIMO",
            "Tecnologia",
            "Tipo",
            "Integraciones / Consume",
            "Cache Adicional al config",
            "Archivo o servicio de donde obtiene informacion para cache",
            "Interaccion con proveedores externos",
            "Metodos que expone",
            "Peso del servicio",
            "Complejidad del servicio",
            "Observacion Discovery",
            "LINK WSDL",
            "LINK CODIGO",
            "Consumen tecnologia deprecada",
            "Peso",
        ]
    )
    ws.append(
        [
            "WSClientes0028",
            "tnd-msa-sp-wsclientes0028",
            "TRIBU",
            "tnd",
            "Bus Omnicanalidad",
            "WS",
            "UMPClientes0020 -> TX067050",
            "",
            "",
            "",
            "ActualizarEmailLocalizacion33",
            "13",
            "Bajo",
            "Validar las descripciones de las tx.",
            "specs",
            "code",
            "",
            "Alta",
        ]
    )
    ws["O2"].hyperlink = (
        "https://dev.azure.com/BancoPichinchaEC/adi-especificaciones-tecnicas/"
        "_git/adi-doc-tecspec-tribu-integracion-apis?path=/sp%20-%20Soporte/"
        "tnd-msa-sp-wsclientes0028"
    )
    ws["P2"].hyperlink = (
        "https://dev.azure.com/BancoPichinchaEC/tpl-bus-omnicanal/"
        "_git/sqb-msa-wsclientes0028"
    )
    wb.save(path)


def _by_id(results, check_id: str):
    return next(r for r in results if r.id == check_id)


def _write_helm_hpa_file(root: Path, file_name: str) -> None:
    """Helm con bloque hpa: (derogado 2026-07) — debe fallar el Check 7.4."""
    helm = root / "helm"
    helm.mkdir(exist_ok=True)
    (helm / file_name).write_text(
        "livenessProbe:\n"
        "  enabled: true\n"
        "readinessProbe:\n"
        "  enabled: true\n"
        "hpa:\n"
        "  minReplicas: 1\n"
        "  maxReplicas: 1\n",
        encoding="utf-8",
    )


def _write_helm_keda_file(root: Path, file_name: str) -> None:
    """Helm con keda: + servicemonitor: (baseline 2026-07) — pasa el Check 7.4."""
    helm = root / "helm"
    helm.mkdir(exist_ok=True)
    (helm / file_name).write_text(
        "livenessProbe:\n"
        "  enabled: true\n"
        "readinessProbe:\n"
        "  enabled: true\n"
        "keda:\n"
        "  enabled: true\n"
        "  minReplicaCount: 1\n"
        "  maxReplicaCount: 1\n"
        "  triggers:\n"
        "    - type: prometheus\n"
        "servicemonitor:\n"
        "  enabled: true\n"
        "  path: '/actuator/prometheus'\n",
        encoding="utf-8",
    )


def _write_helm_env_file(root: Path, file_name: str, value_line: str) -> None:
    helm = root / "helm"
    helm.mkdir(exist_ok=True)
    (helm / file_name).write_text(
        "environment:\n"
        "  - name: \"CCC_WSCLIENTES0006_SKIP_CORRESPONDENCIA_CHANNELS\"\n"
        f"    {value_line}\n",
        encoding="utf-8",
    )


def test_block_1_passes_with_all_layers(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_1(ctx)
    layer_check = _by_id(results, "1.1")
    assert layer_check.status == "pass"


def test_block_1_fails_if_domain_imports_spring(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    bad = root / "src/main/java/com/pichincha/sp/domain/Bad.java"
    bad.write_text(
        "package com.pichincha.sp.domain;\n"
        "import org.springframework.stereotype.Component;\n"
        "public class Bad {}\n",
        encoding="utf-8",
    )
    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_1(ctx)
    check = _by_id(results, "1.2")
    assert check.status == "fail"
    assert check.severity == "high"


def test_block_1_warns_on_legacy_port_layout_for_peer_review(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    legacy_output_port = (
        root
        / "src"
        / "main"
        / "java"
        / "com"
        / "pichincha"
        / "sp"
        / "application"
        / "port"
        / "output"
        / "ClienteOutputPort.java"
    )
    legacy_output_port.parent.mkdir(parents=True, exist_ok=True)
    legacy_output_port.write_text(
        "package com.pichincha.sp.application.port.output;\n"
        "public interface ClienteOutputPort {}\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_1(ctx)
    check = _by_id(results, "1.3b")

    assert check.status == "fail"
    assert check.severity == "medium"
    assert "application/input/port" in check.suggested_fix


def test_block_1_accepts_canonical_peer_review_port_layout(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    canonical_output_port = (
        root
        / "src"
        / "main"
        / "java"
        / "com"
        / "pichincha"
        / "sp"
        / "application"
        / "output"
        / "port"
        / "ClienteOutputPort.java"
    )
    canonical_output_port.parent.mkdir(parents=True, exist_ok=True)
    canonical_output_port.write_text(
        "package com.pichincha.sp.application.output.port;\n"
        "public interface ClienteOutputPort {}\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_1(ctx)
    check = _by_id(results, "1.3b")

    assert check.status == "pass"


def test_block_1_rejects_config_output_port_antipattern(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    config_port = (
        root
        / "src"
        / "main"
        / "java"
        / "com"
        / "pichincha"
        / "sp"
        / "application"
        / "output"
        / "port"
        / "CustomerBasicConfigOutputPort.java"
    )
    config_port.parent.mkdir(parents=True, exist_ok=True)
    config_port.write_text(
        "package com.pichincha.sp.application.output.port;\n"
        "public interface CustomerBasicConfigOutputPort {}\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_1(ctx)
    check = _by_id(results, "1.3c")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "@ConfigurationProperties" in check.suggested_fix


def test_block_7_detects_hardcoded_password(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    yml = root / "src/main/resources/application.yml"
    yml.write_text(
        "spring:\n"
        "  datasource:\n"
        "    password: my-secret-hardcoded\n",
        encoding="utf-8",
    )
    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)
    secret_check = _by_id(results, "7.2")
    assert secret_check.status == "fail"


def test_block_7_passes_with_env_var(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    yml = root / "src/main/resources/application.yml"
    yml.write_text(
        "spring:\n"
        "  datasource:\n"
        "    password: ${CCC_DB_PASSWORD}\n",
        encoding="utf-8",
    )
    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)
    secret_check = _by_id(results, "7.2")
    assert secret_check.status == "pass"


def test_block_7_fails_application_yml_ccc_assignment(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    yml = root / "src/main/resources/application.yml"
    yml.write_text(
        "CCC_WSCLIENTES0006_SKIP_CORRESPONDENCIA_CHANNELS: true\n"
        "app:\n"
        "  skip-correspondence-channels: ${CCC_WSCLIENTES0006_SKIP_CORRESPONDENCIA_CHANNELS}\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)

    check = _by_id(results, "7.2b")
    assert check.status == "fail"
    assert check.severity == "high"
    assert "CCC_WSCLIENTES0006_SKIP_CORRESPONDENCIA_CHANNELS" in check.detail
    assert "helm/dev.yml" in check.suggested_fix


def test_block_7_fails_when_hpa_present(tmp_path: Path) -> None:
    """HPA derogado (2026-07): la presencia del bloque hpa: es HIGH (Check 7.4)."""
    root = _make_migrated(tmp_path)
    (root / "src/main/resources/application.yml").write_text(
        "spring:\n"
        "  datasource:\n"
        "    password: ${CCC_DB_PASSWORD}\n",
        encoding="utf-8",
    )
    for env in ("dev.yml", "test.yml", "prod.yml"):
        _write_helm_hpa_file(root, env)

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)

    hpa_check = _by_id(results, "7.4")
    assert hpa_check.status == "fail"
    assert hpa_check.severity == "high"
    assert "hpa" in hpa_check.detail.lower()
    assert "derogado" in hpa_check.detail.lower()


def test_block_7_passes_with_keda(tmp_path: Path) -> None:
    """Con keda: + servicemonitor: y sin hpa:, el Check 7.4 pasa."""
    root = _make_migrated(tmp_path)
    (root / "src/main/resources/application.yml").write_text(
        "spring:\n"
        "  datasource:\n"
        "    password: ${CCC_DB_PASSWORD}\n",
        encoding="utf-8",
    )
    for env in ("dev.yml", "test.yml", "prod.yml"):
        _write_helm_keda_file(root, env)

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)

    assert _by_id(results, "7.4").status == "pass"
    assert _by_id(results, "7.5d").status == "pass"
    assert _by_id(results, "7.5g").status == "pass"


def test_block_7_fails_helm_env_placeholder_value(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "src/main/resources/application.yml").write_text(
        "app:\n"
        "  skip-correspondence-channels: ${CCC_WSCLIENTES0006_SKIP_CORRESPONDENCIA_CHANNELS}\n",
        encoding="utf-8",
    )
    _write_helm_env_file(
        root,
        "test.yml",
        'value: "<CCC_WSCLIENTES0006_SKIP_CORRESPONDENCIA_CHANNELS_TEST>"',
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)
    check = _by_id(results, "7.5c")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "<CCC_WSCLIENTES0006_SKIP_CORRESPONDENCIA_CHANNELS_TEST>" in check.detail


def test_block_7_fails_helm_generic_placeholder_marker(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "src/main/resources/application.yml").write_text("app:\n  name: test\n", encoding="utf-8")
    helm = root / "helm"
    helm.mkdir(exist_ok=True)
    (helm / "prod.yml").write_text(
        "livenessProbe:\n"
        "  enabled: true\n"
        "readinessProbe:\n"
        "  enabled: true\n"
        "businessConfig:\n"
        "  skipCorrespondencia: <pendiente_validar>\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)
    check = _by_id(results, "7.5c")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "<pendiente_validar>" in check.detail


def test_block_7_fails_helm_env_inline_comment(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "src/main/resources/application.yml").write_text(
        "app:\n"
        "  channel: ${CCC_WSCLIENTES0006_SKIP_CORRESPONDENCIA_CHANNELS}\n",
        encoding="utf-8",
    )
    _write_helm_env_file(root, "dev.yml", 'value: "01,02" # pendiente validar')

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)
    check = _by_id(results, "7.5c")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "comentario inline" in check.detail


def test_block_7_passes_clean_helm_env_value(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "src/main/resources/application.yml").write_text(
        "app:\n"
        "  channel: ${CCC_WSCLIENTES0006_SKIP_CORRESPONDENCIA_CHANNELS}\n",
        encoding="utf-8",
    )
    _write_helm_env_file(root, "prod.yml", 'value: "01,02"')

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)
    check = _by_id(results, "7.5c")

    assert check.status == "pass"


def test_block_7_fails_when_pipeline_namespace_differs_from_catalog(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "src/main/resources/application.yml").write_text(
        "spring:\n  application:\n    name: tnd-msa-sp-wsclientes0026\n",
        encoding="utf-8",
    )
    (root / "catalog-info.yaml").write_text(
        "apiVersion: backstage.io/v1alpha1\n"
        "kind: Component\n"
        "metadata:\n"
        "  name: tpl-middleware\n"
        "  namespace: tnd-middleware\n",
        encoding="utf-8",
    )
    (root / "azure-pipelines.yml").write_text(
        "variables:\n"
        "  KUBERNETES_NAMESPACE: csg-middleware\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)
    check = _by_id(results, "7.6")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "csg-middleware" in check.detail
    assert "tnd-middleware" in check.detail


def test_block_7_fails_when_catalog_namespace_does_not_match_project_prefix(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "src/main/resources/application.yml").write_text(
        "spring:\n  application:\n    name: csg-msa-sp-wsreglas0010\n",
        encoding="utf-8",
    )
    (root / "catalog-info.yaml").write_text(
        "apiVersion: backstage.io/v1alpha1\n"
        "kind: Component\n"
        "metadata:\n"
        "  name: tpl-middleware\n"
        "  namespace: tnd-middleware\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)
    check = _by_id(results, "7.1b")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "csg-middleware" in check.detail
    assert "tnd-middleware" in check.detail


def test_block_7_accepts_catalog_namespace_matching_project_prefix(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "src/main/resources/application.yml").write_text(
        "spring:\n  application:\n    name: csg-msa-sp-wsreglas0010\n",
        encoding="utf-8",
    )
    (root / "catalog-info.yaml").write_text(
        "metadata:\n"
        "  name: tpl-middleware\n"
        "  namespace: csg-middleware\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)
    check = _by_id(results, "7.1b")

    assert check.status == "pass"


def test_block_7_accepts_matching_pipeline_and_catalog_namespace(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "src/main/resources/application.yml").write_text("app:\n  name: test\n", encoding="utf-8")
    (root / "catalog-info.yaml").write_text(
        "metadata:\n"
        "  namespace: tnd-middleware\n",
        encoding="utf-8",
    )
    (root / "azure-pipelines.yml").write_text(
        "variables:\n"
        "- name: KUBERNETES_NAMESPACE\n"
        "  value: tnd-middleware\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_7(ctx)
    check = _by_id(results, "7.6")

    assert check.status == "pass"


def test_block_8_requires_current_spring_boot_plugin(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "build.gradle").write_text(
        "plugins { id 'org.springframework.boot' version '3.5.13' }\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_8(ctx)
    check = _by_id(results, "8.1")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "3.5.14" in check.detail


def test_block_8_passes_current_spring_boot_plugin(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "build.gradle").write_text(
        "plugins { id 'org.springframework.boot' version '3.5.14' }\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_8(ctx)
    check = _by_id(results, "8.1")

    assert check.status == "pass"


def test_block_8_rejects_undertow_dependencies(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "build.gradle").write_text(
        "plugins { id 'org.springframework.boot' version '3.5.14' }\n"
        "dependencies {\n"
        "  def undertowVersion = '2.4.0.RC4'\n"
        "  implementation 'org.springframework.boot:spring-boot-starter-undertow'\n"
        "  implementation \"io.undertow:undertow-core:${undertowVersion}\"\n"
        "}\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_8(ctx)
    check = _by_id(results, "8.2")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "spring-boot-starter-undertow" in check.detail


def test_block_8_accepts_default_embedded_server_without_undertow(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "build.gradle").write_text(
        "plugins { id 'org.springframework.boot' version '3.5.14' }\n"
        "dependencies { implementation 'org.springframework.boot:spring-boot-starter-web' }\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_8(ctx)
    check = _by_id(results, "8.2")

    assert check.status == "pass"


def test_autofix_updates_old_spring_boot_plugin(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    build_gradle = root / "build.gradle"
    build_gradle.write_text(
        "plugins { id 'org.springframework.boot' version '3.5.13' }\n",
        encoding="utf-8",
    )

    def rerun():
        return run_block_8(CheckContext(migrated_path=root, legacy_path=None))

    report = run_autofix_loop(root, rerun)

    assert report.total_applied == 1
    assert "version '3.5.14'" in build_gradle.read_text(encoding="utf-8")


def test_block_0_rejects_bancs_artifacts_for_was(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    resources = root / "src/main/resources"
    resources.mkdir(parents=True, exist_ok=True)
    (resources / "WSClientes0154.wsdl").write_text(
        '<definitions xmlns="http://schemas.xmlsoap.org/wsdl/">'
        '<portType name="P"><operation name="op"/></portType></definitions>',
        encoding="utf-8",
    )
    controller = (
        root
        / "src/main/java/com/pichincha/sp/infrastructure/input/adapter/rest/FooController.java"
    )
    controller.parent.mkdir(parents=True, exist_ok=True)
    controller.write_text(
        "package com.pichincha.sp.infrastructure.input.adapter.rest;\n"
        "import org.springframework.web.bind.annotation.RestController;\n"
        "@RestController class FooController {}\n",
        encoding="utf-8",
    )
    (root / "build.gradle").write_text(
        "dependencies { implementation 'com.pichincha.bnc:lib-bnc-api-client:1.1.0' }\n",
        encoding="utf-8",
    )
    (root / "catalog-info.yaml").write_text(
        "spec:\n  dependsOn:\n    - component:lib-bnc-api-client\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None, source_type="was")
    results = run_block_0(ctx)
    check = _by_id(results, "0.2d")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "BANCS no aplica" in check.detail


def test_block_0_allows_bancs_artifacts_for_iib_with_bancs(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    resources = root / "src/main/resources"
    resources.mkdir(parents=True, exist_ok=True)
    (resources / "WSClientes0006.wsdl").write_text(
        '<definitions xmlns="http://schemas.xmlsoap.org/wsdl/">'
        '<portType name="P"><operation name="op"/></portType></definitions>',
        encoding="utf-8",
    )
    controller = (
        root
        / "src/main/java/com/pichincha/sp/infrastructure/input/adapter/rest/FooController.java"
    )
    controller.parent.mkdir(parents=True, exist_ok=True)
    controller.write_text(
        "package com.pichincha.sp.infrastructure.input.adapter.rest;\n"
        "import org.springframework.web.bind.annotation.RestController;\n"
        "@RestController class FooController {}\n",
        encoding="utf-8",
    )
    (root / "build.gradle").write_text(
        "dependencies { implementation 'com.pichincha.bnc:lib-bnc-api-client:1.1.0' }\n",
        encoding="utf-8",
    )

    ctx = CheckContext(
        migrated_path=root,
        legacy_path=None,
        source_type="iib",
        has_bancs=True,
    )
    results = run_block_0(ctx)
    check = _by_id(results, "0.2d")

    assert check.status == "pass"


def test_block_0_rejects_integrationbus_path_for_was_soap(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    resources = root / "src/main/resources"
    resources.mkdir(parents=True, exist_ok=True)
    (resources / "WSTecnicos0036.wsdl").write_text(
        '<definitions xmlns="http://schemas.xmlsoap.org/wsdl/">'
        '<portType name="P"><operation name="op1"/><operation name="op2"/></portType>'
        "</definitions>",
        encoding="utf-8",
    )
    endpoint = (
        root
        / "src/main/java/com/pichincha/sp/infrastructure/input/adapter/soap/impl/FooEndpoint.java"
    )
    endpoint.parent.mkdir(parents=True, exist_ok=True)
    endpoint.write_text(
        "package com.pichincha.sp.infrastructure.input.adapter.soap.impl;\n"
        "import org.springframework.ws.server.endpoint.annotation.Endpoint;\n"
        "@Endpoint class FooEndpoint {}\n",
        encoding="utf-8",
    )
    config = (
        root
        / "src/main/java/com/pichincha/sp/infrastructure/input/adapter/soap/config/WebServiceConfig.java"
    )
    config.parent.mkdir(parents=True, exist_ok=True)
    config.write_text(
        "package com.pichincha.sp.infrastructure.input.adapter.soap.config;\n"
        "class WebServiceConfig {\n"
        '  String mapping = "/IntegrationBus/soap/*";\n'
        '  String location = "/IntegrationBus/soap/WSTecnicos0036";\n'
        "}\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None, source_type="was")
    results = run_block_0(ctx)
    check = _by_id(results, "0.2f")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "IntegrationBus" in check.detail


def test_block_0_accepts_was_soap_service_path(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    resources = root / "src/main/resources"
    resources.mkdir(parents=True, exist_ok=True)
    (resources / "WSTecnicos0036.wsdl").write_text(
        '<definitions xmlns="http://schemas.xmlsoap.org/wsdl/">'
        '<portType name="P"><operation name="op1"/><operation name="op2"/></portType>'
        "</definitions>",
        encoding="utf-8",
    )
    endpoint = (
        root
        / "src/main/java/com/pichincha/sp/infrastructure/input/adapter/soap/impl/FooEndpoint.java"
    )
    endpoint.parent.mkdir(parents=True, exist_ok=True)
    endpoint.write_text(
        "package com.pichincha.sp.infrastructure.input.adapter.soap.impl;\n"
        "import org.springframework.ws.server.endpoint.annotation.Endpoint;\n"
        "@Endpoint class FooEndpoint {}\n",
        encoding="utf-8",
    )
    config = (
        root
        / "src/main/java/com/pichincha/sp/infrastructure/input/adapter/soap/config/WebServiceConfig.java"
    )
    config.parent.mkdir(parents=True, exist_ok=True)
    config.write_text(
        "package com.pichincha.sp.infrastructure.input.adapter.soap.config;\n"
        "class WebServiceConfig {\n"
        '  String mapping = "/WSTecnicos0036/soap/*";\n'
        '  String location = "/WSTecnicos0036/soap/WSTecnicos0036Request";\n'
        "}\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None, source_type="was")
    results = run_block_0(ctx)
    check = _by_id(results, "0.2f")

    assert check.status == "pass"


def test_block_13_requires_was_connection_test_query(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "build.gradle").write_text(
        "dependencies { implementation 'org.springframework.boot:spring-boot-starter-data-jpa' }\n",
        encoding="utf-8",
    )
    (root / "src/main/resources/application.yml").write_text(
        "spring:\n"
        "  datasource:\n"
        "    hikari:\n"
        "      maximum-pool-size: ${CCC_DB_POOL_MAX}\n"
        "  jpa:\n"
        "    hibernate:\n"
        "      ddl-auto: validate\n"
        "    open-in-view: false\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None, source_type="was")
    results = run_block_13(ctx)
    check = _by_id(results, "13.11")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "connection-test-query: SELECT 1" in check.suggested_fix


def test_block_13_requires_oracle_connection_test_query_from_dual(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "build.gradle").write_text(
        "dependencies { implementation 'org.springframework.boot:spring-boot-starter-data-jpa' }\n",
        encoding="utf-8",
    )
    (root / "src/main/resources/application.yml").write_text(
        "spring:\n"
        "  datasource:\n"
        "    driver-class-name: oracle.jdbc.OracleDriver\n"
        "    hikari:\n"
        "      connection-test-query: SELECT 1\n"
        "  jpa:\n"
        "    hibernate:\n"
        "      ddl-auto: validate\n"
        "    open-in-view: false\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None, source_type="was")
    results = run_block_13(ctx)
    check = _by_id(results, "13.11")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "SELECT 1 from dual" in check.suggested_fix


def test_block_13_accepts_oracle_connection_test_query_from_dual(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "build.gradle").write_text(
        "dependencies { implementation 'org.springframework.boot:spring-boot-starter-data-jpa' }\n",
        encoding="utf-8",
    )
    (root / "src/main/resources/application.yml").write_text(
        "spring:\n"
        "  datasource:\n"
        "    driver-class-name: oracle.jdbc.OracleDriver\n"
        "    hikari:\n"
        "      connection-test-query: SELECT 1 from dual\n"
        "  jpa:\n"
        "    hibernate:\n"
        "      ddl-auto: validate\n"
        "    open-in-view: false\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None, source_type="was")
    results = run_block_13(ctx)
    check = _by_id(results, "13.11")

    assert check.status == "pass"


def test_block_13_accepts_sqlserver_connection_test_query_select_one(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "build.gradle").write_text(
        "dependencies { implementation 'org.springframework.boot:spring-boot-starter-data-jpa' }\n",
        encoding="utf-8",
    )
    (root / "src/main/resources/application.yml").write_text(
        "spring:\n"
        "  datasource:\n"
        "    driver-class-name: com.microsoft.sqlserver.jdbc.SQLServerDriver\n"
        "    hikari:\n"
        "      connection-test-query: SELECT 1\n"
        "  jpa:\n"
        "    hibernate:\n"
        "      ddl-auto: validate\n"
        "    open-in-view: false\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None, source_type="was")
    results = run_block_13(ctx)
    check = _by_id(results, "13.11")

    assert check.status == "pass"


def test_block_13_accepts_was_connection_test_query(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "build.gradle").write_text(
        "dependencies { implementation 'org.springframework.boot:spring-boot-starter-data-jpa' }\n",
        encoding="utf-8",
    )
    (root / "src/main/resources/application.yml").write_text(
        "spring:\n"
        "  datasource:\n"
        "    hikari:\n"
        "      maximum-pool-size: ${CCC_DB_POOL_MAX}\n"
        "      connection-test-query: SELECT 1\n"
        "  jpa:\n"
        "    hibernate:\n"
        "      ddl-auto: validate\n"
        "    open-in-view: false\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None, source_type="was")
    results = run_block_13(ctx)
    check = _by_id(results, "13.11")

    assert check.status == "pass"


def test_block_15_allows_empty_mensaje_negocio_slot(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    mapper = root / "src/main/java/com/pichincha/sp/infrastructure/ErrorMapper.java"
    mapper.write_text(
        "package com.pichincha.sp.infrastructure;\n"
        "class ErrorMapper { void map(Error e) { e.setMensajeNegocio(\"\"); } }\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None)
    results = run_block_15(ctx)
    check = _by_id(results, "15.1")

    assert check.status == "pass"
    assert "aceptado" in check.detail


def _write_migrated_real_mensaje_negocio(root: Path) -> None:
    mapper = root / "src/main/java/com/pichincha/sp/infrastructure/ErrorMapper.java"
    mapper.write_text(
        "package com.pichincha.sp.infrastructure;\n"
        "class ErrorMapper { void map(Error e) { e.setMensajeNegocio(\"Cliente ok\"); } }\n",
        encoding="utf-8",
    )


def _legacy_with_mensaje_negocio(tmp_path: Path, *, populated: bool) -> Path:
    legacy = tmp_path / "legacy" / "svc"
    legacy.mkdir(parents=True, exist_ok=True)
    body = (
        "SET OutputRoot.XMLNSC.error.mensajeNegocio = 'Texto de negocio';\n"
        if populated
        else "SET OutputRoot.XMLNSC.error.codigo = '0';\n"
    )
    (legacy / "flow.esql").write_text(body, encoding="utf-8")
    return tmp_path / "legacy"


def test_block_15_rejects_real_value_when_legacy_did_not_populate(tmp_path: Path) -> None:
    """Valor real en migrado + legacy NO lo poblaba -> HIGH."""
    root = _make_migrated(tmp_path)
    _write_migrated_real_mensaje_negocio(root)
    legacy = _legacy_with_mensaje_negocio(tmp_path, populated=False)

    check = _by_id(run_block_15(CheckContext(migrated_path=root, legacy_path=legacy)), "15.1")

    assert check.status == "fail"
    assert check.severity == "high"


def test_block_15_allows_real_value_when_legacy_populated(tmp_path: Path) -> None:
    """Valor real en migrado + el legacy ya lo poblaba -> PASS (se respeta)."""
    root = _make_migrated(tmp_path)
    _write_migrated_real_mensaje_negocio(root)
    legacy = _legacy_with_mensaje_negocio(tmp_path, populated=True)

    check = _by_id(run_block_15(CheckContext(migrated_path=root, legacy_path=legacy)), "15.1")

    assert check.status == "pass"


def test_block_15_low_when_real_value_and_no_legacy(tmp_path: Path) -> None:
    """Valor real en migrado pero sin legacy para verificar -> LOW (revision manual)."""
    root = _make_migrated(tmp_path)
    _write_migrated_real_mensaje_negocio(root)

    check = _by_id(run_block_15(CheckContext(migrated_path=root, legacy_path=None)), "15.1")

    assert check.status == "fail"
    assert check.severity == "low"


def test_block_21_tx_mapping_passes_when_java_and_yaml_match(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "src/main/resources/application.yml").write_text(
        "bancs:\n  webclients:\n    ws-tx067050:\n      base-url: ${CCC_BANCS_BASE_URL}\n",
        encoding="utf-8",
    )
    adapter = root / "src/main/java/com/pichincha/sp/infrastructure/CustomerBancsAdapter.java"
    adapter.write_text(
        "package com.pichincha.sp.infrastructure;\n"
        "class CustomerBancsAdapter { static final String TRANSACTION_ID = \"067050\"; }\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None, has_bancs=True)
    results = run_block_21(ctx)
    check = _by_id(results, "21.1")

    assert check.status == "pass"


def test_block_21_tx_mapping_fails_when_java_and_yaml_differ(tmp_path: Path) -> None:
    root = _make_migrated(tmp_path)
    (root / "src/main/resources/application.yml").write_text(
        "bancs:\n  webclients:\n    ws-tx067050:\n      base-url: ${CCC_BANCS_BASE_URL}\n",
        encoding="utf-8",
    )
    adapter = root / "src/main/java/com/pichincha/sp/infrastructure/CustomerBancsAdapter.java"
    adapter.write_text(
        "package com.pichincha.sp.infrastructure;\n"
        "class CustomerBancsAdapter { void call() { request.transactionId(\"067051\"); } }\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=root, legacy_path=None, has_bancs=True)
    results = run_block_21(ctx)
    check = _by_id(results, "21.1")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "067050" in check.detail
    assert "067051" in check.detail


def test_block_22_fails_when_discovery_report_is_missing(tmp_path: Path) -> None:
    workspace = tmp_path / "wsclientes0028"
    project = workspace / "destino" / "tnd-msa-sp-wsclientes0028"
    (project / "src/main/java").mkdir(parents=True)
    _make_discovery_workbook(workspace / DISCOVERY_WORKBOOK_NAME)

    ctx = CheckContext(migrated_path=project, legacy_path=None)
    results = run_block_22(ctx)
    check = _by_id(results, "22.1")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "no hay" in check.detail


def test_block_22_fails_when_edge_case_coverage_is_pending(tmp_path: Path) -> None:
    workspace = tmp_path / "wsclientes0028"
    project = workspace / "destino" / "tnd-msa-sp-wsclientes0028"
    (project / "src/main/java").mkdir(parents=True)
    _make_discovery_workbook(workspace / DISCOVERY_WORKBOOK_NAME)
    (workspace / "COMPLEXITY_wsclientes0028.md").write_text(
        "## Discovery / edge cases\n"
        "- Spec path: /sp - Soporte/tnd-msa-sp-wsclientes0028\n"
        "- Code repo: sqb-msa-wsclientes0028\n"
        "DISCOVERY_EDGE_CASES:\n"
        "- edge_cases: tx_description_validation\n"
        "## Discovery edge-case coverage\n"
        "| tx_description_validation | PENDIENTE | <pendiente_validar> |\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=project, legacy_path=None)
    results = run_block_22(ctx)
    check = _by_id(results, "22.2")

    assert check.status == "fail"
    assert check.severity == "high"
    assert "tx_description_validation" in check.detail


def test_block_22_passes_when_edge_case_has_decision_and_file_trace(tmp_path: Path) -> None:
    workspace = tmp_path / "wsclientes0028"
    project = workspace / "destino" / "tnd-msa-sp-wsclientes0028"
    (project / "src/main/java").mkdir(parents=True)
    _make_discovery_workbook(workspace / DISCOVERY_WORKBOOK_NAME)
    (project / "MIGRATION_REPORT.md").write_text(
        "## Discovery / edge cases\n"
        "- Spec path: /sp - Soporte/tnd-msa-sp-wsclientes0028\n"
        "- Code repo: sqb-msa-wsclientes0028\n"
        "DISCOVERY_EDGE_CASES:\n"
        "- edge_cases: tx_description_validation\n"
        "## Discovery edge-case coverage\n"
        "| Codigo | Decision | Implementacion / test |\n"
        "|---|---|---|\n"
        "| tx_description_validation | Decision: implemented | "
        "File: src/test/java/CustomerAdapterTest.java |\n",
        encoding="utf-8",
    )

    ctx = CheckContext(migrated_path=project, legacy_path=None)
    results = run_block_22(ctx)
    report_check = _by_id(results, "22.1")
    coverage_check = _by_id(results, "22.2")

    assert report_check.status == "pass"
    assert coverage_check.status == "pass"
