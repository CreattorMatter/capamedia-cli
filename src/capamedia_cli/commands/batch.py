"""capamedia batch - ejecuta comandos sobre N servicios en paralelo.

Subcomandos:
  - complexity <file>   Analiza complejidad de N servicios legacy (leer de txt)
  - clone      <file>   Clone masivo (legacy + UMPs + TX) con ThreadPool
  - check      <root>   Audita todos los proyectos migrados bajo un path
  - init       <file>   Inicializa N workspaces con .claude/ + CLAUDE.md
  - pipeline   <file>   Ejecuta clone -> init -> fabric -> migrate -> check
  - migrate    <file>   Ejecuta `codex exec` sobre N workspaces listos
  - watch      <root>   Mirador operativo de estados persistentes del batch

Inputs:
  - services.txt: un servicio por linea (ej wsclientes0007). Comentarios con #.
  - --workers N: tamano del pool (default 4)
  - --root <path>: carpeta raiz donde viven los workspaces (default CWD)

Output: tabla consolidada en stdout + archivo `batch-<cmd>-<fecha>.md`.
Opcional: `--xlsx` genera tambien un Excel con la matriz completa.
"""

from __future__ import annotations

import csv
import datetime
import json
import re
import textwrap
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Annotated, Any

import typer
from rich.console import Console
from rich.panel import Panel
from rich.progress import BarColumn, Progress, TextColumn, TimeElapsedColumn
from rich.table import Table

from capamedia_cli.core.batch_state import (
    load_state,
    mark_stage,
    save_state,
    set_result,
    stage_ok,
    state_file,
)
from capamedia_cli.core.engine import (
    Engine,
    EngineInput,
    available_engines,
    engine_from_env,
    select_engine,
)
from capamedia_cli.core.gradle_properties import remove_committed_gradle_java_home
from capamedia_cli.core.scheduler import BatchScheduler
from capamedia_cli.core.self_correction import (
    build_correction_appendix,
    extract_failure_context,
    load_failure_context,
    stash_failure_context,
)

console = Console()

app = typer.Typer(
    help="Procesar N servicios en paralelo (modo batch).",
    no_args_is_help=True,
)

CODEX_REASONING_EFFORTS = {"low", "medium", "high", "xhigh"}
DEFAULT_CODEX_REASONING_EFFORT = "xhigh"


@dataclass
class BatchRow:
    service: str
    status: str  # "ok" | "fail" | "skip" | "wait"
    detail: str
    fields: dict[str, str]


def _normalize_reasoning_effort(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip().lower()
    if not normalized:
        return None
    if normalized not in CODEX_REASONING_EFFORTS:
        valid = " | ".join(sorted(CODEX_REASONING_EFFORTS))
        raise typer.BadParameter(f"reasoning effort invalido: {value}. Usa {valid}")
    return normalized


MIGRATE_FIELD_ORDER = ["engine", "codex", "result", "framework", "build", "check", "seconds", "project"]
PIPELINE_FIELD_ORDER = [
    "engine",
    "clone",
    "init",
    "fabric",
    "codex",
    "result",
    "build",
    "check",
    "seconds",
    "project",
]
WATCH_FIELD_ORDER = [
    "kind",
    "phase",
    "clone",
    "init",
    "fabric",
    "codex",
    "result",
    "build",
    "check",
    "attempts",
    "updated",
    "project",
]

MIGRATE_OUTPUT_SCHEMA: dict[str, Any] = {
    "type": "object",
    "additionalProperties": False,
    "required": [
        "status",
        "summary",
        "framework",
        "build_status",
        "migrated_project",
        "artifacts",
        "notes",
    ],
    "properties": {
        "status": {
            "type": "string",
            "enum": ["ok", "partial", "blocked", "failed"],
        },
        "summary": {"type": "string"},
        "framework": {"type": "string"},
        "build_status": {"type": "string"},
        "migrated_project": {"type": "string"},
        "artifacts": {"type": "array", "items": {"type": "string"}},
        "notes": {"type": "array", "items": {"type": "string"}},
    },
}

FALLBACK_MIGRATE_PROMPT = textwrap.dedent(
    """
    Completa la migracion del servicio legacy a Java 21 + Spring Boot hexagonal siguiendo
    las instrucciones del workspace. Lee primero AGENTS.md y cualquier prompt local del
    proyecto, despues implementa los cambios minimos necesarios dentro de `destino/`.

    Requisitos:
    - trabaja solo dentro del workspace actual;
    - preserva la matriz oficial del proyecto (1 op -> REST/WebFlux, 2+ ops -> SOAP/MVC);
    - valida con build real del proyecto migrado si existe;
    - si falta un prerequisito, no inventes nada: reportalo como `blocked`.

    Tu respuesta final debe ser SOLO un JSON valido segun el schema provisto.
    """
).strip()


def _read_services_file(path: Path, sheet: str | None = None) -> list[str]:
    """Lee servicios desde txt, csv o xlsx.

    - .txt: una linea por servicio. Comentarios con #.
    - .csv: primera columna es el servicio. Header opcional ("servicio" en row 1).
    - .xlsx: hoja `sheet` (default primera). Primera columna = servicio.
      Si la primera celda es literal "servicio", se asume header.
    """
    if not path.exists():
        raise typer.BadParameter(f"no existe: {path}")

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx_services(path, sheet)
    if suffix == ".csv":
        return _read_csv_services(path)

    # Default: .txt
    names: list[str] = []
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        name = line.split("#", 1)[0].strip()
        if name:
            names.append(name)
    return names


def _read_csv_services(path: Path) -> list[str]:
    names: list[str] = []
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if not row or not row[0].strip():
                continue
            cell = row[0].strip()
            if i == 0 and cell.lower() in ("servicio", "service", "name", "nombre"):
                continue  # header
            if cell.startswith("#"):
                continue
            names.append(cell)
    return names


def _read_xlsx_services(path: Path, sheet: str | None) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise typer.BadParameter(
            "openpyxl no disponible. Instalar con: pip install openpyxl"
        ) from e

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    names: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row or row[0] is None:
            continue
        cell = str(row[0]).strip()
        if not cell:
            continue
        if i == 0 and cell.lower() in ("servicio", "service", "name", "nombre"):
            continue  # header
        if cell.startswith("#"):
            continue
        names.append(cell)
    return names


def _write_markdown_report(
    cmd: str, rows: list[BatchRow], workspace: Path, field_order: list[str] | None = None
) -> Path:
    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    dest = workspace / f"batch-{cmd}-{ts}.md"
    lines: list[str] = []
    lines.append(f"# Batch `{cmd}` report\n")
    lines.append(f"Generado: `{datetime.datetime.now().isoformat()}`  \n")
    lines.append(f"Servicios procesados: **{len(rows)}**\n")

    ok = sum(1 for r in rows if r.status == "ok")
    fail = sum(1 for r in rows if r.status == "fail")
    skip = sum(1 for r in rows if r.status == "skip")
    wait = sum(1 for r in rows if r.status == "wait")
    lines.append(f"**OK:** {ok} · **FAIL:** {fail} · **WAIT:** {wait} · **SKIP:** {skip}\n")

    if rows:
        cols = field_order or sorted({k for r in rows for k in r.fields})
        header = ["service", "status", "detail", *cols]
        lines.append("")
        lines.append("| " + " | ".join(header) + " |")
        lines.append("|" + "|".join(["---"] * len(header)) + "|")
        for r in rows:
            row = [r.service, r.status, r.detail[:80]]
            row += [r.fields.get(c, "") for c in cols]
            lines.append("| " + " | ".join(row) + " |")
    dest.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return dest


def _write_csv_report(
    cmd: str, rows: list[BatchRow], workspace: Path, field_order: list[str] | None = None
) -> Path:
    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    dest = workspace / f"batch-{cmd}-{ts}.csv"
    cols = field_order or sorted({k for r in rows for k in r.fields})
    with dest.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["service", "status", "detail", *cols])
        for r in rows:
            w.writerow([r.service, r.status, r.detail, *(r.fields.get(c, "") for c in cols)])
    return dest


def _render_table(cmd: str, rows: list[BatchRow], field_order: list[str] | None = None) -> None:
    cols = field_order or sorted({k for r in rows for k in r.fields})
    table = Table(title=f"Batch {cmd}: {len(rows)} servicios", title_style="bold cyan")
    table.add_column("Servicio", style="cyan")
    table.add_column("Status", style="bold", width=6)
    for c in cols:
        table.add_column(c)
    table.add_column("Detail")
    for r in rows:
        if r.status == "ok":
            status = "[green]OK[/green]"
        elif r.status == "wait":
            status = "[cyan]WAIT[/cyan]"
        elif r.status == "skip":
            status = "[yellow]SKIP[/yellow]"
        else:
            status = "[red]FAIL[/red]"
        vals = [r.fields.get(c, "") for c in cols]
        table.add_row(r.service, status, *vals, r.detail[:50])
    console.print(table)


# -- Helpers compartidos ----------------------------------------------------


def _ensure_batch_runtime_dir(root: Path) -> Path:
    runtime_dir = root / ".capamedia" / "batch"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    return runtime_dir


def _ensure_migrate_schema(root: Path) -> Path:
    runtime_dir = _ensure_batch_runtime_dir(root)
    schema_path = runtime_dir / "codex-batch-migrate.schema.json"
    schema_json = json.dumps(MIGRATE_OUTPUT_SCHEMA, ensure_ascii=True, indent=2) + "\n"
    if not schema_path.exists() or schema_path.read_text(encoding="utf-8") != schema_json:
        schema_path.write_text(schema_json, encoding="utf-8")
    return schema_path


def _has_gradle_build(path: Path) -> bool:
    return (path / "build.gradle").exists() or (path / "build.gradle.kts").exists()


def _find_migrated_project(workspace: Path, service: str) -> Path | None:
    destino = workspace / "destino"
    if not destino.exists():
        return None

    expected_names = (
        f"tnd-msa-sp-{service.lower()}",
        service.lower(),
        service,
    )
    for name in expected_names:
        candidate = destino / name
        if candidate.is_dir() and _has_gradle_build(candidate):
            return candidate

    candidates = sorted(p for p in destino.iterdir() if p.is_dir() and _has_gradle_build(p))
    if not candidates:
        return None

    preferred = [p for p in candidates if service.lower() in p.name.lower()]
    if preferred:
        return preferred[0]
    return candidates[0]


def _find_legacy_root(workspace: Path, service: str | None = None) -> Path | None:
    for candidate in (workspace / "legacy", workspace.parent / "legacy"):
        if candidate.exists() and candidate.is_dir():
            return candidate
    if service:
        from capamedia_cli.core.local_resolver import find_local_legacy

        return find_local_legacy(service, workspace.parent)
    return None


def _has_complexity_report(workspace: Path, service: str) -> bool:
    return (workspace / f"COMPLEXITY_{service}.md").exists()


def _has_init_material(workspace: Path) -> bool:
    return (workspace / ".capamedia" / "config.yaml").exists()


def _load_fabrics_metadata(workspace: Path) -> dict[str, Any] | None:
    from capamedia_cli.commands.fabrics import load_fabrics_metadata

    data = load_fabrics_metadata(workspace)
    return data if isinstance(data, dict) else None


def _has_fabrics_material(workspace: Path) -> bool:
    data = _load_fabrics_metadata(workspace)
    if not data:
        return False
    project_path = _as_text(data.get("project_path"))
    project_name = _as_text(data.get("project_name"))
    return bool(project_path or project_name)


def _find_project_from_fabrics_metadata(workspace: Path) -> Path | None:
    data = _load_fabrics_metadata(workspace)
    if not data:
        return None
    project_path = _as_text(data.get("project_path")).strip()
    if project_path:
        candidate = Path(project_path)
        if candidate.is_dir() and _has_gradle_build(candidate):
            return candidate
    project_name = _as_text(data.get("project_name")).strip()
    if project_name:
        candidate = workspace / "destino" / project_name
        if candidate.is_dir() and _has_gradle_build(candidate):
            return candidate
    return None


def _format_ts(raw: str) -> str:
    value = raw.strip()
    if not value:
        return ""
    return value.replace("T", " ")[:19]


def _read_state_snapshot(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return data if isinstance(data, dict) else None


def _watch_stage_value(state: dict[str, Any], stage: str, field_key: str, default: str = "") -> str:
    stage_data = state.get("stages", {}).get(stage, {})
    if not isinstance(stage_data, dict):
        return default
    fields = stage_data.get("fields", {})
    if isinstance(fields, dict) and fields.get(field_key) is not None:
        return _as_text(fields.get(field_key))
    return _as_text(stage_data.get("status"), default)


def _watch_phase(run_kind: str, state: dict[str, Any]) -> str:
    order = ["clone", "init", "fabric", "migrate", "check"] if run_kind == "pipeline" else ["migrate", "check"]
    stages = state.get("stages", {})
    for stage in order:
        stage_data = stages.get(stage, {})
        if not isinstance(stage_data, dict):
            return stage
        status = _as_text(stage_data.get("status")).strip().lower()
        if status == "fail":
            return f"{stage}:fail"
        if not status:
            return stage
        if status != "ok":
            return f"{stage}:{status}"
    result_status = _as_text(state.get("result", {}).get("status")).strip().lower()
    if result_status == "ok":
        return "done"
    return result_status or "pending"


def _watch_row_for_service(workspace: Path, service: str, run_kind: str) -> BatchRow:
    state = _read_state_snapshot(state_file(workspace, run_kind))
    fabrics = _load_fabrics_metadata(workspace) or {}
    project_hint = _as_text(fabrics.get("project_name")) or _as_text(
        state.get("result", {}).get("fields", {}).get("project") if state else ""
    )

    if state is None:
        return BatchRow(
            service,
            "wait",
            _as_text(fabrics.get("detail"), "sin estado batch"),
            {
                "kind": run_kind,
                "phase": "no_state",
                "clone": "",
                "init": "",
                "fabric": _as_text(fabrics.get("status"), "-"),
                "codex": "",
                "result": "",
                "build": "",
                "check": "",
                "attempts": "0",
                "updated": "",
                "project": project_hint,
            },
        )

    result = state.get("result", {}) if isinstance(state.get("result"), dict) else {}
    attempts = 0
    for stage_data in state.get("stages", {}).values():
        if isinstance(stage_data, dict):
            attempts += int(stage_data.get("attempts", 0) or 0)

    fields = {
        "kind": run_kind,
        "phase": _watch_phase(run_kind, state),
        "clone": _watch_stage_value(state, "clone", "clone"),
        "init": _watch_stage_value(state, "init", "init"),
        "fabric": _watch_stage_value(state, "fabric", "fabric", _as_text(fabrics.get("status"), "")),
        "codex": _watch_stage_value(state, "migrate", "codex"),
        "result": _as_text(result.get("fields", {}).get("result")) if isinstance(result.get("fields"), dict) else _as_text(result.get("status")),
        "build": _watch_stage_value(state, "migrate", "build"),
        "check": _watch_stage_value(state, "check", "check"),
        "attempts": str(attempts),
        "updated": _format_ts(_as_text(state.get("updated_at"))),
        "project": _as_text(
            (result.get("fields", {}) if isinstance(result.get("fields"), dict) else {}).get("project"),
            project_hint,
        ),
    }

    failed_stage_detail = ""
    for stage_name in ("clone", "init", "fabric", "migrate", "check"):
        stage_data = state.get("stages", {}).get(stage_name, {})
        if isinstance(stage_data, dict) and _as_text(stage_data.get("status")).lower() == "fail":
            failed_stage_detail = _as_text(stage_data.get("detail"))
            break

    result_status = _as_text(result.get("status")).lower()
    if result_status == "ok":
        status = "ok"
    elif failed_stage_detail:
        status = "fail"
    else:
        status = "wait"

    detail = _as_text(result.get("detail")) or failed_stage_detail or _as_text(fabrics.get("detail"), "en progreso")
    return BatchRow(service, status, detail, fields)


def _collect_watch_rows(root: Path, services: list[str], kind: str) -> list[BatchRow]:
    rows: list[BatchRow] = []
    for service in services:
        workspace = root / service
        if kind == "pipeline":
            rows.append(_watch_row_for_service(workspace, service, "pipeline"))
            continue
        if kind == "migrate":
            rows.append(_watch_row_for_service(workspace, service, "migrate"))
            continue
        pipeline_state = state_file(workspace, "pipeline")
        migrate_state = state_file(workspace, "migrate")
        if pipeline_state.exists():
            rows.append(_watch_row_for_service(workspace, service, "pipeline"))
        elif migrate_state.exists():
            rows.append(_watch_row_for_service(workspace, service, "migrate"))
        else:
            rows.append(_watch_row_for_service(workspace, service, "pipeline"))
    rows.sort(key=lambda r: r.service)
    return rows


def _hydrate_fields(base: dict[str, str], saved: dict[str, Any] | None) -> dict[str, str]:
    fields = dict(base)
    if not isinstance(saved, dict):
        return fields
    for key, value in saved.items():
        fields[str(key)] = _as_text(value)
    return fields


def _load_migrate_prompt(workspace: Path, prompt_file: Path | None = None) -> str:
    candidate = prompt_file or (workspace / ".codex" / "prompts" / "migrate.md")
    if candidate.exists():
        return candidate.read_text(encoding="utf-8").strip()
    return FALLBACK_MIGRATE_PROMPT


def _build_batch_migrate_prompt(
    service: str,
    workspace: Path,
    migrated_project: Path,
    prompt_body: str,
) -> str:
    from capamedia_cli.core.catalog_injector import (
        contains_catalog_block,
        detect_relevant_tx,
        format_for_prompt,
        load_catalogs,
    )

    legacy_root = _find_legacy_root(workspace, service)
    legacy_hint = str(legacy_root) if legacy_root else "(no encontrado)"

    # Inyeccion de catalogos oficiales: evitar que la AI alucine TX-BANCS,
    # codigos de backend o reglas de error. Si el prompt_body ya trae el
    # bloque (viene del FABRICS_PROMPT_<svc>.md), no duplicamos.
    if contains_catalog_block(prompt_body):
        catalog_block = ""
    else:
        tx_codes = detect_relevant_tx(workspace, service)
        snapshot = load_catalogs(workspace)
        catalog_block = format_for_prompt(snapshot, relevant_tx=tx_codes)

    base = textwrap.dedent(
        f"""
        Servicio objetivo: {service}
        Workspace root: {workspace}
        Legacy root: {legacy_hint}
        Proyecto migrado esperado: {migrated_project}

        Ejecuta la migracion de forma no interactiva en este workspace ya preparado.
        Antes de editar:
        1. Lee `AGENTS.md` si existe.
        2. Usa el prompt base incluido abajo como fuente principal del workflow.
        3. Si falta un prerequisito importante, devolve `status=blocked` con detalle concreto.

        Requisitos operativos:
        - Trabaja solo dentro de este workspace.
        - Corre al menos un build real del proyecto migrado si el proyecto existe.
        - Nunca agregues ni mantengas `org.gradle.java.home` en `gradle.properties`;
          las rutas locales de JDK rompen Azure DevOps/Linux. Si necesitas Java 21,
          usa `JAVA_HOME` del entorno del proceso.
        - No incluyas Markdown ni explicaciones fuera del JSON final.
        - La respuesta final debe ser SOLO el objeto JSON pedido por el schema.
        - Inclui siempre todas las claves del schema: status, summary, framework,
          build_status, migrated_project, artifacts y notes. Si no aplica, usa
          string vacio o lista vacia.

        Prompt base del proyecto:
        ---
        {prompt_body}
        ---
        """
    ).strip()

    if catalog_block:
        return base + "\n\n" + catalog_block.rstrip()
    return base


def _strip_code_fence(raw: str) -> str:
    raw = raw.strip()
    if not raw.startswith("```"):
        return raw
    lines = raw.splitlines()
    if len(lines) >= 2 and lines[-1].strip() == "```":
        return "\n".join(lines[1:-1]).strip()
    return raw


def _read_structured_message(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}

    raw = _strip_code_fence(path.read_text(encoding="utf-8").strip())
    if not raw:
        return {}

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        start = raw.find("{")
        end = raw.rfind("}")
        if start == -1 or end == -1 or end <= start:
            return {}
        try:
            data = json.loads(raw[start : end + 1])
        except json.JSONDecodeError:
            return {}

    return data if isinstance(data, dict) else {}


def _summarize_engine_output(stderr: str, stdout: str, default: str) -> str:
    text = (stderr.strip() or stdout.strip()).strip()
    if not text:
        return default

    error_match = re.search(
        r'"code"\s*:\s*"([^"]+)".*?"message"\s*:\s*"([^"]+)"',
        text,
        re.DOTALL,
    )
    if error_match:
        return f"{error_match.group(1)}: {error_match.group(2)}"

    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("ERROR:"):
            return stripped[:160]
    return text.splitlines()[0][:160]


def _as_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, str):
        return value
    return str(value)


def _normalize_build_status(value: str) -> str:
    lowered = value.strip().lower()
    if lowered in {"green", "ok", "passed", "pass", "success", "successful"}:
        return "green"
    if lowered in {"red", "fail", "failed", "error"}:
        return "red"
    if lowered in {"", "unknown", "not_run", "not run"}:
        return "unknown"
    return value.strip() or "unknown"


def _run_batch_check(service: str, migrated_project: Path, legacy_root: Path | None) -> dict[str, str]:
    from capamedia_cli.commands.check import _write_report
    from capamedia_cli.core.checklist_rules import CheckContext, run_all_blocks

    ctx = CheckContext(migrated_path=migrated_project, legacy_path=legacy_root)
    results = run_all_blocks(ctx)
    report_path = _write_report(service, results, migrated_project, legacy_root)

    high = sum(1 for r in results if r.status == "fail" and r.severity == "high")
    medium = sum(1 for r in results if r.status == "fail" and r.severity == "medium")
    low = sum(1 for r in results if r.status == "fail" and r.severity == "low")

    if high > 0:
        verdict = "BLOCKED_BY_HIGH"
    elif medium > 0:
        verdict = "READY_WITH_FOLLOW_UP"
    else:
        verdict = "READY_TO_MERGE"

    return {
        "verdict": verdict,
        "HIGH": str(high),
        "MEDIUM": str(medium),
        "LOW": str(low),
        "report": str(report_path),
    }


def _process_migrate_service(
    service: str,
    root: Path,
    schema_path: Path,
    *,
    engine: Engine,
    model: str | None,
    prompt_file: Path | None,
    timeout_minutes: int,
    run_check: bool,
    unsafe: bool,
    reasoning_effort: str | None = None,
    resume: bool = False,
    scheduler: BatchScheduler | None = None,
    stream_output: bool = False,
) -> BatchRow:
    workspace = root / service
    return _process_migrate_workspace(
        service,
        workspace,
        schema_path,
        engine=engine,
        model=model,
        prompt_file=prompt_file,
        timeout_minutes=timeout_minutes,
        run_check=run_check,
        unsafe=unsafe,
        reasoning_effort=reasoning_effort,
        resume=resume,
        scheduler=scheduler,
        stream_output=stream_output,
    )


def _process_migrate_workspace(
    service: str,
    workspace: Path,
    schema_path: Path,
    *,
    engine: Engine,
    model: str | None,
    prompt_file: Path | None,
    timeout_minutes: int,
    run_check: bool,
    unsafe: bool,
    reasoning_effort: str | None = None,
    resume: bool = False,
    scheduler: BatchScheduler | None = None,
    stream_output: bool = False,
) -> BatchRow:
    if not workspace.exists():
        return BatchRow(service, "fail", f"workspace no existe: {workspace}", {})

    state = load_state(workspace, "migrate", service, reset=not resume)
    state_result = state.get("result", {}) if isinstance(state.get("result"), dict) else {}
    base_fields = {
        "engine": engine.name,
        "codex": "pending",
        "result": "pending",
        "framework": "?",
        "build": "unknown",
        "check": "not_run" if run_check else "skip",
        "seconds": "0.0",
        "project": "?",
    }

    fabrics_meta = _load_fabrics_metadata(workspace)
    if not fabrics_meta:
        fields = _hydrate_fields(base_fields, state_result.get("fields"))
        detail = (
            "no hay evidencia de Fabrics en .capamedia/fabrics.json "
            "(corre `capamedia fabrics generate` o `batch pipeline` antes de migrate)"
        )
        set_result(state, status="fail", detail=detail, fields=fields)
        mark_stage(state, "migrate", status="fail", detail=detail, fields=fields)
        save_state(workspace, "migrate", state)
        return BatchRow(service, "fail", detail, fields)

    migrated_project = _find_project_from_fabrics_metadata(workspace) or _find_migrated_project(workspace, service)
    if migrated_project is None:
        fields = _hydrate_fields(base_fields, state_result.get("fields"))
        detail = "no se encontro proyecto migrado en destino/ pese a tener metadata de Fabrics"
        set_result(state, status="fail", detail=detail, fields=fields)
        mark_stage(state, "migrate", status="fail", detail=detail, fields=fields)
        save_state(workspace, "migrate", state)
        return BatchRow(
            service,
            "fail",
            detail,
            fields,
        )

    remove_committed_gradle_java_home(migrated_project)

    if resume and state_result.get("status") == "ok" and stage_ok(state, "migrate"):
        check_is_done = (not run_check) or stage_ok(state, "check")
        if check_is_done:
            fields = _hydrate_fields(base_fields, state_result.get("fields"))
            fields["project"] = migrated_project.name
            return BatchRow(
                service,
                "ok",
                _as_text(state_result.get("detail"), "already completed (resume)") or "already completed (resume)",
                fields,
            )

    prompt_body = _load_migrate_prompt(workspace, prompt_file)
    prompt = _build_batch_migrate_prompt(service, workspace, migrated_project, prompt_body)

    # Self-correction: si es un retry y hay un FailureContext cacheado en el
    # state del servicio, inyectamos un appendix con el error especifico.
    failure_ctx = load_failure_context(state)
    if failure_ctx is not None:
        prompt = build_correction_appendix(failure_ctx, prompt)

    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = workspace / ".capamedia" / "batch-migrate"
    run_dir.mkdir(parents=True, exist_ok=True)
    prompt_path = run_dir / f"{engine.name}-prompt-{ts}.md"
    output_path = run_dir / f"{engine.name}-last-message-{ts}.json"
    stdout_path = run_dir / f"{engine.name}-stdout-{ts}.log"
    stderr_path = run_dir / f"{engine.name}-stderr-{ts}.log"
    prompt_path.write_text(prompt + "\n", encoding="utf-8")

    skip_migrate = resume and stage_ok(state, "migrate")
    project_name = migrated_project.name
    result_status = "ok"
    build_status = "unknown"
    summary = ""
    elapsed = 0.0
    if skip_migrate:
        saved_fields = {}
        stage_data = state.get("stages", {}).get("migrate", {})
        if isinstance(stage_data, dict):
            saved_fields = stage_data.get("fields", {})
            summary = _as_text(stage_data.get("detail"), "migrate already completed")
        project_name = _as_text(saved_fields.get("project"), migrated_project.name) or migrated_project.name
        result_status = _as_text(saved_fields.get("result"), "ok").lower()
        build_status = _normalize_build_status(_as_text(saved_fields.get("build"), "green"))
    else:
        einput = EngineInput(
            workspace=workspace,
            prompt=prompt,
            schema_path=schema_path,
            output_path=output_path,
            timeout_seconds=timeout_minutes * 60,
            model=model,
            reasoning_effort=reasoning_effort,
            unsafe=unsafe,
            stream_output=stream_output,
        )
        if scheduler is not None:
            scheduler.acquire(service)
        try:
            try:
                eres = engine.run_headless(einput)
                # Si detecto rate limit, pauso global y reintento una vez
                if eres.rate_limited and scheduler is not None:
                    scheduler.handle_rate_limit(service, eres.retry_after_seconds)
                    eres = engine.run_headless(einput)
            finally:
                remove_committed_gradle_java_home(migrated_project)
        finally:
            if scheduler is not None:
                scheduler.release(service)

        stdout_text = eres.stdout
        stderr_text = eres.stderr
        returncode = eres.exit_code

        if eres.failure_reason and eres.exit_code == 127:
            # binario no encontrado
            fields = _hydrate_fields(base_fields, state_result.get("fields"))
            fields["project"] = migrated_project.name
            detail = f"engine {engine.name}: {eres.failure_reason}"
            set_result(state, status="fail", detail=detail, fields=fields)
            mark_stage(state, "migrate", status="fail", detail=detail, fields=fields)
            save_state(workspace, "migrate", state)
            return BatchRow(service, "fail", detail, fields)

        if eres.failure_reason and eres.exit_code == 124:
            # timeout
            stdout_path.write_text(stdout_text, encoding="utf-8")
            stderr_path.write_text(stderr_text, encoding="utf-8")
            fields = _hydrate_fields(base_fields, state_result.get("fields"))
            fields.update(
                {
                    "engine": engine.name,
                    "codex": "timeout",
                    "result": "failed",
                    "framework": "?",
                    "build": "unknown",
                    "check": "not_run" if run_check else "skip",
                    "seconds": f"{eres.duration_seconds:.1f}",
                    "project": migrated_project.name,
                }
            )
            set_result(state, status="fail", detail=eres.failure_reason, fields=fields)
            mark_stage(state, "migrate", status="fail", detail=eres.failure_reason, fields=fields)
            save_state(workspace, "migrate", state)
            return BatchRow(service, "fail", eres.failure_reason, fields)

        elapsed = eres.duration_seconds
        stdout_path.write_text(stdout_text, encoding="utf-8")
        stderr_path.write_text(stderr_text, encoding="utf-8")

        payload = _read_structured_message(output_path)
        result_status = _as_text(payload.get("status"), "ok" if returncode == 0 else "failed").lower()
        summary = _as_text(payload.get("summary")).strip()
        if not summary:
            summary = _summarize_engine_output(
                stderr_text,
                stdout_text,
                f"{engine.name} termino sin resumen estructurado",
            )
        framework = _as_text(payload.get("framework"), "?").strip() or "?"
        build_status = _normalize_build_status(_as_text(payload.get("build_status"), "unknown"))
        reported_project = _as_text(payload.get("migrated_project"), str(migrated_project)).strip()
        project_name = Path(reported_project).name if reported_project else migrated_project.name
        migrate_fields = {
            "engine": engine.name,
            "codex": "ok" if returncode == 0 else f"exit_{returncode}",
            "result": result_status,
            "framework": framework,
            "build": build_status,
            "check": "not_run" if run_check else "skip",
            "seconds": f"{elapsed:.1f}",
            "project": project_name,
        }
        if eres.rate_limited:
            migrate_fields["rate_limited"] = "yes"
        migrate_stage_status = "ok" if returncode == 0 and result_status == "ok" and build_status == "green" else "fail"
        mark_stage(state, "migrate", status=migrate_stage_status, detail=summary.splitlines()[0][:160], fields=migrate_fields)
        save_state(workspace, "migrate", state)

    saved_migrate_fields = {}
    migrate_stage_data = state.get("stages", {}).get("migrate", {})
    if isinstance(migrate_stage_data, dict):
        saved_migrate_fields = migrate_stage_data.get("fields", {})

    fields = _hydrate_fields(base_fields, saved_migrate_fields or state_result.get("fields"))
    fields["project"] = project_name
    if not skip_migrate and elapsed > 0:
        fields["seconds"] = f"{elapsed:.1f}"

    check_status = fields.get("check", "not_run" if run_check else "skip")
    if run_check:
        if resume and stage_ok(state, "check"):
            stage_data = state.get("stages", {}).get("check", {})
            if isinstance(stage_data, dict):
                saved_check_fields = stage_data.get("fields", {})
                fields["check"] = _as_text(saved_check_fields.get("check"), _as_text(stage_data.get("detail"), "READY_TO_MERGE"))
            check_status = fields["check"]
        else:
            try:
                check_result = _run_batch_check(service, migrated_project, _find_legacy_root(workspace, service))
                check_status = check_result["verdict"]
                fields["check"] = check_status
                mark_stage(
                    state,
                    "check",
                    status="ok" if not check_status.startswith("BLOCKED_BY_HIGH") else "fail",
                    detail=check_status,
                    fields={"check": check_status, "report": check_result["report"]},
                )
            except Exception as exc:
                check_status = f"CHECK_ERROR: {type(exc).__name__}"
                fields["check"] = check_status
                mark_stage(state, "check", status="fail", detail=check_status, fields={"check": check_status})
            save_state(workspace, "migrate", state)

    row_status = "ok"
    if fields.get("codex") not in {"ok"} or result_status != "ok" or build_status != "green" or check_status.startswith("BLOCKED_BY_HIGH") or check_status.startswith("CHECK_ERROR"):
        row_status = "fail"

    detail = (summary or _as_text(state_result.get("detail")) or "migrate completed").splitlines()[0][:160]
    set_result(state, status=row_status, detail=detail, fields=fields)
    save_state(workspace, "migrate", state)
    return BatchRow(service, row_status, detail, fields)


def _process_pipeline_service(
    service: str,
    root: Path,
    schema_path: Path,
    *,
    harnesses: list[str],
    artifact_token: str | None,
    namespace: str,
    group_id: str,
    engine: Engine,
    model: str | None,
    prompt_file: Path | None,
    timeout_minutes: int,
    skip_tx: bool,
    shallow: bool,
    skip_check: bool,
    unsafe: bool,
    reasoning_effort: str | None = None,
    resume: bool = False,
    scheduler: BatchScheduler | None = None,
) -> BatchRow:
    from capamedia_cli.commands.clone import clone_service
    from capamedia_cli.commands.fabrics import generate, inspect_fabrics_workspace
    from capamedia_cli.commands.init import scaffold_project

    workspace = root / service
    workspace.mkdir(parents=True, exist_ok=True)
    state = load_state(workspace, "pipeline", service, reset=not resume)

    fields = {
        "engine": engine.name,
        "clone": "pending",
        "init": "pending",
        "fabric": "pending",
        "codex": "pending",
        "result": "pending",
        "build": "pending",
        "check": "pending",
        "seconds": "0.0",
        "project": "?",
    }

    started = time.perf_counter()

    if not (resume and stage_ok(state, "clone") and _find_legacy_root(workspace, service) and _has_complexity_report(workspace, service)):
        try:
            clone_service(service, workspace=workspace, shallow=shallow, skip_tx=skip_tx)
            fields["clone"] = "ok"
            mark_stage(state, "clone", status="ok", detail="clone completed", fields={"clone": "ok"})
            save_state(workspace, "pipeline", state)
        except typer.Exit:
            fields["clone"] = "fail"
            fields["seconds"] = f"{time.perf_counter() - started:.1f}"
            mark_stage(state, "clone", status="fail", detail="clone failed", fields={"clone": "fail"})
            set_result(state, status="fail", detail="clone failed", fields=fields)
            save_state(workspace, "pipeline", state)
            return BatchRow(service, "fail", "clone failed", fields)
        except Exception as exc:
            fields["clone"] = "fail"
            fields["seconds"] = f"{time.perf_counter() - started:.1f}"
            detail = f"clone error: {type(exc).__name__}: {exc}"
            mark_stage(state, "clone", status="fail", detail=detail, fields={"clone": "fail"})
            set_result(state, status="fail", detail=detail, fields=fields)
            save_state(workspace, "pipeline", state)
            return BatchRow(service, "fail", detail, fields)
    else:
        fields["clone"] = "ok"

    if not (resume and stage_ok(state, "init") and _has_init_material(workspace)):
        try:
            scaffold_project(
                target_dir=workspace,
                service_name=service,
                harnesses=harnesses,
                artifact_token=artifact_token,
            )
            fields["init"] = "ok"
            mark_stage(state, "init", status="ok", detail="init completed", fields={"init": "ok"})
            save_state(workspace, "pipeline", state)
        except Exception as exc:
            fields["init"] = "fail"
            fields["seconds"] = f"{time.perf_counter() - started:.1f}"
            detail = f"init error: {type(exc).__name__}: {exc}"
            mark_stage(state, "init", status="fail", detail=detail, fields={"init": "fail"})
            set_result(state, status="fail", detail=detail, fields=fields)
            save_state(workspace, "pipeline", state)
            return BatchRow(service, "fail", detail, fields)
    else:
        fields["init"] = "ok"

    if not (resume and stage_ok(state, "fabric") and _has_fabrics_material(workspace)):
        fabrics_ready = inspect_fabrics_workspace(workspace)
        if fabrics_ready["status"] != "ok":
            fields["fabric"] = "fail"
            fields["seconds"] = f"{time.perf_counter() - started:.1f}"
            detail = f"fabrics preflight failed: {fabrics_ready['detail']}"
            mark_stage(state, "fabric", status="fail", detail=detail, fields={"fabric": "fail"})
            set_result(state, status="fail", detail=detail, fields=fields)
            save_state(workspace, "pipeline", state)
            return BatchRow(service, "fail", detail, fields)
        try:
            generate(
                service_name=service,
                workspace=workspace,
                namespace=namespace,
                group_id=group_id,
                dry_run=False,
            )
            fabrics_meta = _load_fabrics_metadata(workspace)
            if not fabrics_meta:
                raise RuntimeError("Fabrics no dejo metadata en .capamedia/fabrics.json")
            fabric_status = _as_text(fabrics_meta.get("status"), "ok") or "ok"
            fabric_detail = _as_text(fabrics_meta.get("detail"), "fabric completed") or "fabric completed"
            fields["fabric"] = fabric_status
            fields["project"] = _as_text(fabrics_meta.get("project_name"), fields["project"])
            mark_stage(
                state,
                "fabric",
                status="ok",
                detail=fabric_detail,
                fields={"fabric": fabric_status, "project": fields["project"]},
            )
            save_state(workspace, "pipeline", state)
        except typer.Exit:
            fields["fabric"] = "fail"
            fields["seconds"] = f"{time.perf_counter() - started:.1f}"
            mark_stage(state, "fabric", status="fail", detail="fabric failed", fields={"fabric": "fail"})
            set_result(state, status="fail", detail="fabric failed", fields=fields)
            save_state(workspace, "pipeline", state)
            return BatchRow(service, "fail", "fabric failed", fields)
        except Exception as exc:
            fields["fabric"] = "fail"
            fields["seconds"] = f"{time.perf_counter() - started:.1f}"
            detail = f"fabric error: {type(exc).__name__}: {exc}"
            mark_stage(state, "fabric", status="fail", detail=detail, fields={"fabric": "fail"})
            set_result(state, status="fail", detail=detail, fields=fields)
            save_state(workspace, "pipeline", state)
            return BatchRow(service, "fail", detail, fields)
    else:
        fabrics_meta = _load_fabrics_metadata(workspace)
        fields["fabric"] = _as_text((fabrics_meta or {}).get("status"), "ok")
        fields["project"] = _as_text((fabrics_meta or {}).get("project_name"), fields["project"])

    migrate_row = _process_migrate_service(
        service,
        root,
        schema_path,
        engine=engine,
        model=model,
        reasoning_effort=reasoning_effort,
        prompt_file=prompt_file,
        timeout_minutes=timeout_minutes,
        run_check=not skip_check,
        unsafe=unsafe,
        resume=resume,
        scheduler=scheduler,
    )
    fields.update(migrate_row.fields)
    fields["seconds"] = f"{time.perf_counter() - started:.1f}"
    mark_stage(state, "migrate", status="ok" if migrate_row.status == "ok" else "fail", detail=migrate_row.detail, fields=migrate_row.fields)
    if not skip_check:
        mark_stage(
            state,
            "check",
            status="ok" if migrate_row.fields.get("check", "").startswith("READY") else "fail",
            detail=migrate_row.fields.get("check", ""),
            fields={"check": migrate_row.fields.get("check", "")},
        )
    set_result(state, status=migrate_row.status, detail=migrate_row.detail, fields=fields)
    save_state(workspace, "pipeline", state)
    return BatchRow(service, migrate_row.status, migrate_row.detail, fields)


def _run_service_with_retries(
    service: str,
    runner,
    *,
    retries: int,
    workspace_resolver=None,
    project_resolver=None,
    run_kind: str = "migrate",
) -> BatchRow:
    """Ejecuta runner(service, attempt) hasta que devuelva OK o se agoten retries.

    Si se pasa `workspace_resolver(service) -> Path`, entre un attempt fallido y
    el siguiente se extrae un FailureContext y se guarda en el state del servicio
    para que el runner lo pueda leer y armar un prompt de self-correction.

    `project_resolver(service) -> Path | None` es opcional y se usa para leer el
    CHECKLIST_*.md del proyecto migrado. Si no se provee, solo se extraen build
    errors + tails.
    """
    last_row: BatchRow | None = None
    for attempt in range(retries + 1):
        if attempt > 0 and workspace_resolver is not None:
            try:
                workspace = workspace_resolver(service)
            except Exception:
                workspace = None
            if workspace is not None and workspace.exists():
                migrated_project = None
                if project_resolver is not None:
                    try:
                        migrated_project = project_resolver(service, workspace)
                    except Exception:
                        migrated_project = None
                try:
                    state = load_state(workspace, run_kind, service, reset=False)
                    ctx = extract_failure_context(workspace, migrated_project, state)
                    stash_failure_context(state, ctx)
                    save_state(workspace, run_kind, state)
                except Exception:
                    # No queremos romper el retry loop si el parseo falla:
                    # fallback al comportamiento previo (prompt a secas).
                    pass
        row = runner(service, attempt)
        if row.status == "ok":
            return row
        last_row = row
    assert last_row is not None
    return last_row


def _ensure_legacy_available(service: str, root: Path, shallow: bool) -> tuple[Path | None, str]:
    """Resuelve el legacy del servicio: local first, fallback a Azure multi-project.

    Retorna (path_legacy, source_msg). Si path es None, source_msg indica el motivo de error.
    """
    from capamedia_cli.commands.clone import _resolve_azure_repo
    from capamedia_cli.core.local_resolver import find_local_legacy

    # 1. Buscar local primero
    capa_media_root = root.parent
    local = find_local_legacy(service, capa_media_root)
    if local:
        return (local, "local")

    # 2. Fallback a Azure DevOps multi-proyecto / multi-patron
    svc_workspace = root / service
    svc_workspace.mkdir(parents=True, exist_ok=True)
    legacy_dest, project_key, repo_name = _resolve_azure_repo(service, svc_workspace, shallow)
    if legacy_dest is None:
        return (None, "not local + no Azure project found")
    return (legacy_dest, f"azure: {project_key}/{repo_name}")


# Backward compat
def _ensure_legacy_cloned(service: str, root: Path, shallow: bool) -> tuple[bool, str]:
    """Wrapper legacy para tests viejos."""
    path, msg = _ensure_legacy_available(service, root, shallow)
    return (path is not None, msg)


# -- Subcommands ------------------------------------------------------------


@app.command("complexity")
def batch_complexity(
    file: Annotated[
        Path,
        typer.Option("--from", "-f", help="Archivo .txt / .csv / .xlsx con un servicio por linea"),
    ],
    sheet: Annotated[
        str | None,
        typer.Option("--sheet", help="Hoja del .xlsx (default: primera)"),
    ] = None,
    workers: Annotated[int, typer.Option("--workers", "-w")] = 4,
    root: Annotated[
        Path | None,
        typer.Option("--root", help="Carpeta raiz donde clonar workspaces (default: CWD)"),
    ] = None,
    shallow: Annotated[bool, typer.Option("--shallow")] = True,
    csv_out: Annotated[bool, typer.Option("--csv", help="Ademas del .md, genera un .csv")] = False,
) -> None:
    """Analiza complejidad de N servicios en paralelo."""
    from capamedia_cli.core.legacy_analyzer import analyze_legacy

    services = _read_services_file(file, sheet=sheet)
    ws = (root or Path.cwd()).resolve()

    console.print(
        Panel.fit(
            f"[bold]batch complexity[/bold]\n"
            f"Servicios: {len(services)} · Workers: {workers} · Root: {ws}",
            border_style="cyan",
        )
    )

    rows: list[BatchRow] = []

    def process(service: str) -> BatchRow:
        legacy_root, msg = _ensure_legacy_available(service, ws, shallow)
        if legacy_root is None:
            return BatchRow(service, "fail", msg, {})
        try:
            analysis = analyze_legacy(legacy_root, service_name=service, umps_root=None)
        except Exception as e:
            return BatchRow(service, "fail", f"analyze error: {e}", {})

        # Detectar dominios distintos via UMPs
        from capamedia_cli.core.domain_mapping import domains_for_umps

        ump_names = [u.name for u in analysis.umps]
        domains = domains_for_umps(ump_names)
        domain_str = "+".join(d.pascal for d in domains) if domains else "-"

        return BatchRow(
            service,
            "ok",
            f"source={msg}",
            {
                "tipo": analysis.source_kind.upper(),
                "ops": str(analysis.wsdl.operation_count) if analysis.wsdl else "?",
                "framework": analysis.framework_recommendation.upper() or "?",
                "umps": str(len(analysis.umps)),
                "dominios": domain_str,
                "bd": "SI" if analysis.has_database else "NO",
                "complejidad": analysis.complexity.upper(),
            },
        )

    with Progress(
        TextColumn("[bold cyan]{task.description}"),
        BarColumn(),
        TextColumn("{task.completed}/{task.total}"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Analizando", total=len(services))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            for future in as_completed(pool.submit(process, s) for s in services):
                rows.append(future.result())
                progress.advance(task)

    rows.sort(key=lambda r: r.service)
    field_order = ["tipo", "ops", "framework", "umps", "dominios", "bd", "complejidad"]

    _render_table("complexity", rows, field_order)
    md = _write_markdown_report("complexity", rows, ws, field_order)
    console.print(f"\n[bold]Reporte:[/bold] [cyan]{md}[/cyan]")
    if csv_out:
        csvf = _write_csv_report("complexity", rows, ws, field_order)
        console.print(f"[bold]CSV:[/bold] [cyan]{csvf}[/cyan]")


@app.command("clone")
def batch_clone(
    file: Annotated[
        Path, typer.Option("--from", "-f", help="Archivo .txt / .csv / .xlsx con servicios")
    ],
    sheet: Annotated[
        str | None, typer.Option("--sheet", help="Hoja del .xlsx (default: primera)")
    ] = None,
    workers: Annotated[int, typer.Option("--workers", "-w")] = 4,
    root: Annotated[Path | None, typer.Option("--root")] = None,
    shallow: Annotated[bool, typer.Option("--shallow")] = False,
    init: Annotated[
        bool,
        typer.Option(
            "--init",
            help="v0.23.4: al terminar cada clone, correr `capamedia init` automaticamente. "
            "Default harness: claude (configurable con --init-ai). Equivale a correr "
            "`batch clone --from <f>` + `batch init --from <f>` en una sola linea.",
        ),
    ] = False,
    init_ai: Annotated[
        str,
        typer.Option(
            "--init-ai",
            help="Harness AI para el init automatico. Default: claude. "
            "CSV permitido (ej. claude,codex). Solo aplica si --init esta activado.",
        ),
    ] = "claude",
) -> None:
    """Clone masivo (legacy + UMPs + TX) en paralelo.

    Con `--init`: tambien corre init del harness (default Claude) al finalizar
    cada clone. Equivale al flujo single `capamedia clone <svc> --init` de v0.23.0
    pero para N servicios.
    """
    from capamedia_cli.commands.clone import clone_service

    services = _read_services_file(file, sheet=sheet)
    ws = (root or Path.cwd()).resolve()
    header_suffix = f" + init ({init_ai})" if init else ""
    console.print(
        Panel.fit(
            f"[bold]batch clone{header_suffix}[/bold]\n"
            f"Servicios: {len(services)} · Workers: {workers} · Root: {ws}",
            border_style="cyan",
        )
    )

    rows: list[BatchRow] = []

    # Resolver harnesses UNA vez antes del pool (validar que el nombre sea bueno)
    harnesses: list[str] = []
    if init:
        from capamedia_cli.adapters import resolve_harnesses

        harnesses = resolve_harnesses(init_ai)

    def process(service: str) -> BatchRow:
        svc_ws = ws / service
        svc_ws.mkdir(parents=True, exist_ok=True)
        try:
            clone_service(service, workspace=svc_ws, shallow=shallow, skip_tx=False)
        except typer.Exit:
            return BatchRow(service, "fail", "clone failed (typer.Exit)", {})
        except Exception as e:
            return BatchRow(service, "fail", f"{type(e).__name__}: {e}", {})

        extra: dict[str, str] = {"workspace": str(svc_ws.name)}
        if init:
            try:
                from capamedia_cli.commands.init import scaffold_project

                scaffold_project(
                    target_dir=svc_ws,
                    service_name=service,
                    harnesses=harnesses,
                    artifact_token=None,
                )
                extra["init"] = "ok"
                extra["harness"] = init_ai
            except Exception as e:
                # clone OK pero init fallo: reportar mixed success
                extra["init"] = f"fail: {type(e).__name__}"
                return BatchRow(service, "partial", str(svc_ws), extra)

        return BatchRow(service, "ok", str(svc_ws), extra)

    with Progress(
        TextColumn("[bold cyan]{task.description}"),
        BarColumn(),
        TextColumn("{task.completed}/{task.total}"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        label = "Clone + init" if init else "Clonando"
        task = progress.add_task(label, total=len(services))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            for future in as_completed(pool.submit(process, s) for s in services):
                rows.append(future.result())
                progress.advance(task)

    rows.sort(key=lambda r: r.service)
    cols = ["workspace", "init", "harness"] if init else ["workspace"]
    _render_table("clone", rows, cols)
    md = _write_markdown_report("clone", rows, ws, cols)
    console.print(f"\n[bold]Reporte:[/bold] [cyan]{md}[/cyan]")


@app.command("check")
def batch_check(
    root: Annotated[Path, typer.Argument(help="Directorio raiz donde viven los workspaces")],
    glob_pattern: Annotated[
        str,
        typer.Option(
            "--glob",
            help="Glob relativo desde root para localizar proyectos migrados. Ej: '*/destino/*'",
        ),
    ] = "*/destino/*",
    workers: Annotated[int, typer.Option("--workers", "-w")] = 4,
) -> None:
    """Audita checklist sobre N proyectos migrados en paralelo."""
    from capamedia_cli.core.checklist_rules import CheckContext, run_all_blocks

    root = root.resolve()
    projects = [p for p in root.glob(glob_pattern) if (p / "build.gradle").exists()]
    if not projects:
        console.print(f"[yellow]No se encontraron proyectos con build.gradle bajo {root / glob_pattern}[/yellow]")
        raise typer.Exit(0)

    console.print(
        Panel.fit(
            f"[bold]batch check[/bold]\n"
            f"Proyectos: {len(projects)} · Workers: {workers} · Root: {root}",
            border_style="cyan",
        )
    )

    rows: list[BatchRow] = []

    def process(proj: Path) -> BatchRow:
        # Auto-descubrir legacy hermano
        legacy_path = None
        for candidate in (proj.parent.parent / "legacy", proj.parent / "legacy"):
            if candidate.exists():
                legacy_path = candidate
                break
        ctx = CheckContext(migrated_path=proj, legacy_path=legacy_path)
        try:
            results = run_all_blocks(ctx)
        except Exception as e:
            return BatchRow(proj.name, "fail", f"check error: {e}", {})

        high = sum(1 for r in results if r.status == "fail" and r.severity == "high")
        medium = sum(1 for r in results if r.status == "fail" and r.severity == "medium")
        low = sum(1 for r in results if r.status == "fail" and r.severity == "low")
        pass_ = sum(1 for r in results if r.status == "pass")

        if high > 0:
            verdict = "BLOCKED_BY_HIGH"
        elif medium > 0:
            verdict = "READY_WITH_FOLLOW_UP"
        else:
            verdict = "READY_TO_MERGE"

        return BatchRow(
            proj.name,
            "ok" if high == 0 else "fail",
            verdict,
            {
                "verdict": verdict,
                "pass": str(pass_),
                "HIGH": str(high),
                "MEDIUM": str(medium),
                "LOW": str(low),
            },
        )

    with Progress(
        TextColumn("[bold cyan]{task.description}"),
        BarColumn(),
        TextColumn("{task.completed}/{task.total}"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Auditando", total=len(projects))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            for future in as_completed(pool.submit(process, p) for p in projects):
                rows.append(future.result())
                progress.advance(task)

    rows.sort(key=lambda r: r.service)
    field_order = ["verdict", "pass", "HIGH", "MEDIUM", "LOW"]
    _render_table("check", rows, field_order)
    md = _write_markdown_report("check", rows, root, field_order)
    console.print(f"\n[bold]Reporte:[/bold] [cyan]{md}[/cyan]")


@app.command("review")
def batch_review(
    root: Annotated[
        Path | None,
        typer.Argument(
            help="Carpeta raiz donde viven los workspaces de servicios. Default: CWD.",
        ),
    ] = None,
    file: Annotated[
        Path | None,
        typer.Option(
            "--from", "-f",
            help="Archivo .txt/.csv/.xlsx con nombres de servicios. Si se omite, "
            "autodetecta subcarpetas de root que tengan destino/ + legacy/.",
        ),
    ] = None,
    sheet: Annotated[
        str | None, typer.Option("--sheet", help="Hoja del .xlsx")
    ] = None,
    workers: Annotated[int, typer.Option("--workers", "-w")] = 2,
    skip_official: Annotated[
        bool,
        typer.Option(
            "--skip-official",
            help="Saltear fase 4 (validador oficial). Util si no tenes venv Python ok.",
        ),
    ] = False,
) -> None:
    """Review masivo sobre N workspaces (v0.23.4).

    Parado en una carpeta que contiene varios workspaces (cada uno con
    `destino/` + `legacy/`), corre `capamedia review` para todos en paralelo
    y produce una tabla consolidada de veredictos.

    Input (cualquiera):
      - `--from <file>` para especificar nombres explicitos
      - Sin argumentos: autodetecta subcarpetas de CWD con `destino/` + `legacy/`

    Ideal para:
      - Cerrar un sprint donde migraste varios servicios — revisarlos de una
      - Validar que los N migrados del equipo esten listos para PR
    """
    from capamedia_cli.core.checklist_rules import CheckContext, run_all_blocks

    root_path = (root or Path.cwd()).resolve()

    # 1. Resolver lista de workspaces
    if file:
        service_names = _read_services_file(file, sheet=sheet)
        workspaces = [root_path / svc for svc in service_names]
    else:
        # Autodetect: buscar subcarpetas con destino/ + legacy/ o solo destino/
        workspaces = []
        for subdir in sorted(root_path.iterdir()):
            if not subdir.is_dir() or subdir.name.startswith("."):
                continue
            if (subdir / "destino").is_dir():
                workspaces.append(subdir)
        if not workspaces:
            console.print(
                f"[red]Error:[/red] no se encontraron workspaces bajo {root_path}. "
                "Pasa `--from <archivo>` con nombres explicitos, o parate en una "
                "carpeta que contenga subcarpetas de servicios con `destino/`."
            )
            raise typer.Exit(code=2)

    console.print(
        Panel.fit(
            f"[bold]batch review[/bold]\n"
            f"Workspaces: {len(workspaces)} · Workers: {workers} · Root: {root_path}\n"
            f"Validator oficial: {'SKIP' if skip_official else 'ON'}",
            border_style="cyan",
        )
    )

    rows: list[BatchRow] = []

    def process(svc_ws: Path) -> BatchRow:
        svc_name = svc_ws.name
        # Localizar el project_path (destino/<unico-subdir>) y legacy_path
        destino = svc_ws / "destino"
        if not destino.is_dir():
            return BatchRow(svc_name, "fail", "sin carpeta destino/", {})
        destino_subs = [p for p in destino.iterdir() if p.is_dir() and not p.name.startswith(".")]
        if len(destino_subs) != 1:
            return BatchRow(
                svc_name, "fail",
                f"destino/ tiene {len(destino_subs)} subdirs (esperaba 1)",
                {},
            )
        project_path = destino_subs[0]

        legacy_root = svc_ws / "legacy"
        legacy_path = None
        if legacy_root.is_dir():
            legacy_subs = [p for p in legacy_root.iterdir() if p.is_dir() and not p.name.startswith(".")]
            if len(legacy_subs) == 1:
                legacy_path = legacy_subs[0]

        # Popular source_type / has_bancs via detector
        source_type = ""
        has_bancs = False
        if legacy_path and legacy_path.is_dir():
            try:
                from capamedia_cli.core.legacy_analyzer import (
                    detect_bancs_connection,
                    detect_source_kind,
                )
                source_type = detect_source_kind(legacy_path, svc_name)
                has_bancs, _ = detect_bancs_connection(legacy_path)
            except Exception:
                pass

        ctx = CheckContext(
            migrated_path=project_path,
            legacy_path=legacy_path,
            source_type=source_type,
            has_bancs=has_bancs,
        )

        try:
            results = run_all_blocks(ctx)
        except Exception as e:
            return BatchRow(svc_name, "fail", f"checklist error: {e}", {})

        high = sum(1 for r in results if r.status == "fail" and r.severity == "high")
        medium = sum(1 for r in results if r.status == "fail" and r.severity == "medium")
        low = sum(1 for r in results if r.status == "fail" and r.severity == "low")
        pass_ = sum(1 for r in results if r.status == "pass")

        if high > 0:
            verdict = "BLOCKED_BY_HIGH"
            status = "fail"
        elif medium > 0:
            verdict = "READY_WITH_FOLLOW_UP"
            status = "ok"
        else:
            verdict = "READY_TO_MERGE"
            status = "ok"

        return BatchRow(
            svc_name, status, verdict,
            {
                "verdict": verdict,
                "source": source_type.upper() or "?",
                "pass": str(pass_),
                "HIGH": str(high),
                "MEDIUM": str(medium),
                "LOW": str(low),
            },
        )

    with Progress(
        TextColumn("[bold cyan]{task.description}"),
        BarColumn(),
        TextColumn("{task.completed}/{task.total}"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Review", total=len(workspaces))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            for future in as_completed(pool.submit(process, w) for w in workspaces):
                rows.append(future.result())
                progress.advance(task)

    rows.sort(key=lambda r: r.service)
    field_order = ["verdict", "source", "pass", "HIGH", "MEDIUM", "LOW"]
    _render_table("review", rows, field_order)
    md = _write_markdown_report("review", rows, root_path, field_order)
    console.print(f"\n[bold]Reporte:[/bold] [cyan]{md}[/cyan]")


@app.command("init")
def batch_init(
    file: Annotated[
        Path, typer.Option("--from", "-f", help="Archivo .txt / .csv / .xlsx con servicios")
    ],
    sheet: Annotated[
        str | None, typer.Option("--sheet", help="Hoja del .xlsx (default: primera)")
    ] = None,
    ai: Annotated[str, typer.Option("--ai", help="Harness(es) CSV o 'all'")] = "claude",
    workers: Annotated[int, typer.Option("--workers", "-w")] = 4,
    root: Annotated[Path | None, typer.Option("--root")] = None,
) -> None:
    """Inicializa N workspaces con .claude/ + CLAUDE.md + .mcp.json."""
    from capamedia_cli.adapters import resolve_harnesses
    from capamedia_cli.commands.init import scaffold_project

    services = _read_services_file(file, sheet=sheet)
    ws = (root or Path.cwd()).resolve()
    harnesses = resolve_harnesses(ai)
    console.print(
        Panel.fit(
            f"[bold]batch init[/bold]\n"
            f"Servicios: {len(services)} · Harness: {ai} · Workers: {workers}",
            border_style="cyan",
        )
    )

    rows: list[BatchRow] = []

    def process(service: str) -> BatchRow:
        svc_ws = ws / service
        svc_ws.mkdir(parents=True, exist_ok=True)
        try:
            scaffold_project(
                target_dir=svc_ws,
                service_name=service,
                harnesses=harnesses,
                artifact_token=None,
            )
            return BatchRow(service, "ok", "inicializado", {"harness": ai})
        except Exception as e:
            return BatchRow(service, "fail", f"{type(e).__name__}: {e}", {})

    with Progress(
        TextColumn("[bold cyan]{task.description}"),
        BarColumn(),
        TextColumn("{task.completed}/{task.total}"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Inicializando", total=len(services))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            for future in as_completed(pool.submit(process, s) for s in services):
                rows.append(future.result())
                progress.advance(task)

    rows.sort(key=lambda r: r.service)
    _render_table("init", rows, ["harness"])
    md = _write_markdown_report("init", rows, ws, ["harness"])
    console.print(f"\n[bold]Reporte:[/bold] [cyan]{md}[/cyan]")


@app.command("watch")
def batch_watch(
    root: Annotated[Path, typer.Argument(help="Directorio raiz donde viven los workspaces")],
    file: Annotated[
        Path | None,
        typer.Option("--from", "-f", help="Archivo opcional .txt / .csv / .xlsx con servicios"),
    ] = None,
    sheet: Annotated[
        str | None, typer.Option("--sheet", help="Hoja del .xlsx (default: primera)")
    ] = None,
    kind: Annotated[
        str,
        typer.Option("--kind", help="Que estado mirar: auto | pipeline | migrate"),
    ] = "auto",
    failures_only: Annotated[
        bool, typer.Option("--failures-only", help="Mostrar solo servicios fallidos")
    ] = False,
    follow: Annotated[
        bool, typer.Option("--follow", help="Refresca periodicamente hasta Ctrl+C")
    ] = False,
    interval_seconds: Annotated[
        int, typer.Option("--interval-seconds", help="Intervalo de refresh en modo --follow (legacy)")
    ] = 15,
    rich_mode: Annotated[
        bool | None,
        typer.Option(
            "--rich/--plain",
            help="Dashboard rich con barras (default: TTY si, pipe no). --plain fuerza tabla plana.",
        ),
    ] = None,
    refresh_seconds: Annotated[
        float,
        typer.Option("--refresh-seconds", help="Refresh del dashboard rich (default 2s)"),
    ] = 2.0,
    engine: Annotated[
        str, typer.Option("--engine", help="Etiqueta del engine mostrado en el footer")
    ] = "codex",
) -> None:
    """Mirador operativo de lotes batch usando el estado persistente por servicio."""
    root = root.resolve()
    if kind not in {"auto", "pipeline", "migrate"}:
        raise typer.BadParameter("kind debe ser uno de: auto, pipeline, migrate")

    if file is not None:
        services = _read_services_file(file, sheet=sheet)
    else:
        services = sorted(
            p.name
            for p in root.iterdir()
            if p.is_dir()
            and not p.name.startswith(".")
            and any((p / marker).exists() for marker in (".capamedia", "legacy", "destino"))
        )

    if not services:
        console.print(f"[yellow]No se encontraron servicios bajo {root}[/yellow]")
        raise typer.Exit(0)

    use_rich = console.is_terminal if rich_mode is None else rich_mode

    if use_rich:
        from rich.live import Live

        from capamedia_cli.core.dashboard import Dashboard, render_rich

        dashboard = Dashboard(root, services=services, kind=kind, engine=engine)

        def _build_renderable():
            snaps = dashboard.snapshot()
            if failures_only:
                snaps = [s for s in snaps if s.status == "failed"]
            agg = dashboard.aggregate(snaps)
            return snaps, agg, render_rich(
                snaps,
                agg,
                title=f"Batch migration: {root}",
            )

        snaps, agg, renderable = _build_renderable()

        if not follow:
            console.print(renderable)
            md = _write_markdown_report(
                f"watch-{kind}", _collect_watch_rows(root, services, kind), root, WATCH_FIELD_ORDER
            )
            console.print(f"\n[bold]Reporte:[/bold] [cyan]{md}[/cyan]")
            return

        refresh_every = max(0.5, float(refresh_seconds))
        try:
            with Live(renderable, console=console, refresh_per_second=max(1, int(1 / refresh_every))) as live:
                while True:
                    time.sleep(refresh_every)
                    snaps, agg, renderable = _build_renderable()
                    live.update(renderable)
                    if agg.total > 0 and (agg.done + agg.failed) == agg.total:
                        break
        except KeyboardInterrupt:
            console.print("\n[yellow]watch cancelado por el usuario[/yellow]")

        # Reporte final resumido.
        _print_rich_summary(snaps, agg)
        return

    def render_snapshot() -> list[BatchRow]:
        rows = _collect_watch_rows(root, services, kind)
        if failures_only:
            rows = [row for row in rows if row.status == "fail"]
        _render_table("watch", rows, WATCH_FIELD_ORDER)
        return rows

    if follow:
        try:
            while True:
                console.clear()
                console.print(
                    Panel.fit(
                        f"[bold]batch watch[/bold]\n"
                        f"Servicios: {len(services)} · Root: {root}\n"
                        f"Kind: {kind} · Failures only: {'SI' if failures_only else 'NO'} · Every: {interval_seconds}s",
                        border_style="cyan",
                    )
                )
                render_snapshot()
                time.sleep(interval_seconds)
        except KeyboardInterrupt:
            console.print("\n[yellow]watch cancelado por el usuario[/yellow]")
        return

    console.print(
        Panel.fit(
            f"[bold]batch watch[/bold]\n"
            f"Servicios: {len(services)} · Root: {root}\n"
            f"Kind: {kind} · Failures only: {'SI' if failures_only else 'NO'}",
            border_style="cyan",
        )
    )
    rows = render_snapshot()
    md = _write_markdown_report(f"watch-{kind}", rows, root, WATCH_FIELD_ORDER)
    console.print(f"\n[bold]Reporte:[/bold] [cyan]{md}[/cyan]")


def _print_rich_summary(snaps, aggregate) -> None:
    """Imprime tabla final tras `watch --rich --follow`."""
    table = Table(title="Batch watch: resumen final", title_style="bold cyan")
    table.add_column("Servicio", style="cyan")
    table.add_column("Fase")
    table.add_column("Status")
    table.add_column("Intentos", justify="right")
    for snap in snaps:
        style = "red" if snap.status == "failed" else "green" if snap.status == "done" else "cyan"
        table.add_row(
            snap.name,
            snap.phase,
            f"[{style}]{snap.status}[/{style}]",
            str(snap.attempts),
        )
    console.print(table)
    console.print(
        f"[bold]Totales:[/bold] done={aggregate.done} failed={aggregate.failed} "
        f"running={aggregate.running} queued={aggregate.queued} "
        f"success_rate={aggregate.success_rate * 100:.0f}%"
    )


@app.command("pipeline")
def batch_pipeline(
    file: Annotated[
        Path,
        typer.Option("--from", "-f", help="Archivo .txt / .csv / .xlsx con servicios"),
    ],
    namespace: Annotated[
        str,
        typer.Option(
            "--namespace",
            "-n",
            help="Namespace del catalogo para Fabrics (obligatorio en modo batch)",
        ),
    ],
    sheet: Annotated[
        str | None, typer.Option("--sheet", help="Hoja del .xlsx (default: primera)")
    ] = None,
    ai: Annotated[
        str,
        typer.Option(
            "--ai",
            help="Harness(es) CSV para el scaffold. `codex` se agrega automaticamente si falta.",
        ),
    ] = "codex",
    workers: Annotated[int, typer.Option("--workers", "-w")] = 2,
    root: Annotated[Path | None, typer.Option("--root")] = None,
    group_id: Annotated[str, typer.Option("--group-id")] = "com.pichincha.sp",
    artifact_token: Annotated[
        str | None,
        typer.Option("--artifact-token", help="Override del token para renderizar .mcp.json"),
    ] = None,
    engine_name: Annotated[
        str,
        typer.Option(
            "--engine",
            help="Engine AI headless: claude | codex | auto (default codex; auto prioriza Claude si esta disponible)",
        ),
    ] = "codex",
    claude_bin: Annotated[
        str, typer.Option("--claude-bin", help="Binario de Claude Code CLI")
    ] = "claude",
    codex_bin: Annotated[str, typer.Option("--codex-bin", help="Binario de Codex CLI")] = "codex",
    model: Annotated[
        str | None,
        typer.Option(
            "--model",
            help=(
                "Override del modelo del engine seleccionado. Para Codex, default desde "
                "~/.codex/config.toml o GPT-5.5 si el workspace fue generado por capamedia init."
            ),
        ),
    ] = None,
    reasoning_effort: Annotated[
        str | None,
        typer.Option(
            "--reasoning-effort",
            help=(
                "Override de razonamiento para Codex CLI: low | medium | high | xhigh. "
                "Default recomendado: xhigh."
            ),
        ),
    ] = DEFAULT_CODEX_REASONING_EFFORT,
    prompt_file: Annotated[
        Path | None,
        typer.Option(
            "--prompt-file",
            help="Prompt base alternativo para todos los servicios. Default: <workspace>/.codex/prompts/migrate.md",
        ),
    ] = None,
    timeout_minutes: Annotated[
        int, typer.Option("--timeout-minutes", help="Timeout maximo por servicio para migrate")
    ] = 90,
    shallow: Annotated[
        bool, typer.Option("--shallow", help="Clone superficial para legacy/UMPs/TX")
    ] = False,
    skip_tx: Annotated[
        bool, typer.Option("--skip-tx", help="No clonar repos individuales de TX")
    ] = False,
    skip_check: Annotated[
        bool, typer.Option("--skip-check", help="No ejecutar checklist post-migracion")
    ] = False,
    resume: Annotated[
        bool, typer.Option("--resume", help="Reanuda servicios saltando etapas ya exitosas")
    ] = False,
    retries: Annotated[
        int, typer.Option("--retries", help="Reintentos adicionales por servicio")
    ] = 0,
    unsafe: Annotated[
        bool,
        typer.Option(
            "--unsafe",
            help="Usa permisos full para el engine (bypass de sandbox/approvals)",
        ),
    ] = False,
    max_services_per_window: Annotated[
        int,
        typer.Option(
            "--max-services-per-window",
            help="Throttle proactivo: servicios maximo por ventana de suscripcion. 0=off.",
        ),
    ] = 0,
    window_hours: Annotated[
        float,
        typer.Option(
            "--window-hours",
            help="Duracion de la ventana de suscripcion (Claude Max = 5h).",
        ),
    ] = 5.0,
) -> None:
    """Ejecuta el pipeline completo por servicio en paralelo."""
    from capamedia_cli.adapters import resolve_harnesses

    services = _read_services_file(file, sheet=sheet)
    ws = (root or Path.cwd()).resolve()

    if prompt_file is not None:
        prompt_file = prompt_file.resolve()
        if not prompt_file.exists():
            raise typer.BadParameter(f"prompt no existe: {prompt_file}")

    # Engine selection (env -> flag)
    env_pref = engine_from_env()
    eff_engine_name = env_pref or engine_name
    try:
        engine = select_engine(eff_engine_name, claude_bin=claude_bin, codex_bin=codex_bin)
    except (RuntimeError, ValueError) as exc:
        raise typer.BadParameter(str(exc)) from None

    scheduler = (
        BatchScheduler(
            services_per_window=max_services_per_window,
            window_seconds=window_hours * 3600,
            on_event=lambda m: console.print(f"[yellow]{m}[/yellow]"),
        )
        if max_services_per_window > 0
        else None
    )

    harnesses = resolve_harnesses(ai)
    if engine.name not in harnesses:
        harnesses.append(engine.name)

    schema_path = _ensure_migrate_schema(ws)
    eff_reasoning = _normalize_reasoning_effort(reasoning_effort if engine.name == "codex" else None)

    console.print(
        Panel.fit(
            f"[bold]batch pipeline[/bold]\n"
            f"Servicios: {len(services)} · Workers: {workers} · Root: {ws}\n"
            f"Harnesses: {', '.join(harnesses)} · Namespace: {namespace}\n"
            f"Engine: [green]{engine.name}[/green] ({engine.subscription_type}) · "
            f"Model: {model or '(default)'} · Reasoning: {eff_reasoning or '(engine default)'}\n"
            f"Resume: {'SI' if resume else 'NO'} · Retries: {retries} · "
            f"Window: {max_services_per_window or 'off'}/{window_hours}h",
            border_style="cyan",
        )
    )

    rows: list[BatchRow] = []

    with Progress(
        TextColumn("[bold cyan]{task.description}"),
        BarColumn(),
        TextColumn("{task.completed}/{task.total}"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Pipeline", total=len(services))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = [
                pool.submit(
                    _run_service_with_retries,
                    service,
                    lambda svc, attempt: _process_pipeline_service(
                        svc,
                        ws,
                        schema_path,
                        harnesses=harnesses,
                        artifact_token=artifact_token,
                        namespace=namespace,
                        group_id=group_id,
                        engine=engine,
                        model=model,
                        reasoning_effort=eff_reasoning,
                        prompt_file=prompt_file,
                        timeout_minutes=timeout_minutes,
                        skip_tx=skip_tx,
                        shallow=shallow,
                        skip_check=skip_check,
                        unsafe=unsafe,
                        resume=resume or attempt > 0,
                        scheduler=scheduler,
                    ),
                    retries=retries,
                )
                for service in services
            ]
            for future in as_completed(futures):
                rows.append(future.result())
                progress.advance(task)

    rows.sort(key=lambda r: r.service)
    _render_table("pipeline", rows, PIPELINE_FIELD_ORDER)
    md = _write_markdown_report("pipeline", rows, ws, PIPELINE_FIELD_ORDER)
    console.print(f"\n[bold]Reporte:[/bold] [cyan]{md}[/cyan]")


@app.command("migrate")
def batch_migrate(
    file: Annotated[
        Path,
        typer.Option("--from", "-f", help="Archivo .txt / .csv / .xlsx con servicios"),
    ],
    sheet: Annotated[
        str | None, typer.Option("--sheet", help="Hoja del .xlsx (default: primera)")
    ] = None,
    workers: Annotated[int, typer.Option("--workers", "-w")] = 2,
    root: Annotated[Path | None, typer.Option("--root")] = None,
    engine_name: Annotated[
        str,
        typer.Option(
            "--engine",
            help="Engine AI headless: claude | codex | auto (default codex; auto prioriza Claude si esta disponible)",
        ),
    ] = "codex",
    claude_bin: Annotated[
        str, typer.Option("--claude-bin", help="Binario de Claude Code CLI")
    ] = "claude",
    codex_bin: Annotated[str, typer.Option("--codex-bin", help="Binario de Codex CLI")] = "codex",
    model: Annotated[
        str | None,
        typer.Option(
            "--model",
            help=(
                "Override del modelo del engine seleccionado. Para Codex, default desde "
                "~/.codex/config.toml o GPT-5.5 si el workspace fue generado por capamedia init."
            ),
        ),
    ] = None,
    reasoning_effort: Annotated[
        str | None,
        typer.Option(
            "--reasoning-effort",
            help=(
                "Override de razonamiento para Codex CLI: low | medium | high | xhigh. "
                "Default recomendado: xhigh."
            ),
        ),
    ] = DEFAULT_CODEX_REASONING_EFFORT,
    prompt_file: Annotated[
        Path | None,
        typer.Option(
            "--prompt-file",
            help="Prompt base alternativo para todos los servicios. Default: <workspace>/.codex/prompts/migrate.md",
        ),
    ] = None,
    timeout_minutes: Annotated[
        int, typer.Option("--timeout-minutes", help="Timeout maximo por servicio")
    ] = 90,
    skip_check: Annotated[
        bool, typer.Option("--skip-check", help="No ejecutar checklist post-migracion")
    ] = False,
    resume: Annotated[
        bool, typer.Option("--resume", help="Reanuda servicios ya migrados exitosamente")
    ] = False,
    retries: Annotated[
        int, typer.Option("--retries", help="Reintentos adicionales por servicio")
    ] = 0,
    unsafe: Annotated[
        bool,
        typer.Option(
            "--unsafe",
            help="Usa permisos full para el engine (bypass de sandbox/approvals)",
        ),
    ] = False,
    max_services_per_window: Annotated[
        int,
        typer.Option(
            "--max-services-per-window",
            help="Throttle proactivo: servicios maximo por ventana de suscripcion. 0=off.",
        ),
    ] = 0,
    window_hours: Annotated[
        float,
        typer.Option(
            "--window-hours",
            help="Duracion de la ventana de suscripcion (Claude Max = 5h).",
        ),
    ] = 5.0,
) -> None:
    """Ejecuta el engine headless en paralelo sobre workspaces ya preparados."""
    services = _read_services_file(file, sheet=sheet)
    ws = (root or Path.cwd()).resolve()

    if prompt_file is not None:
        prompt_file = prompt_file.resolve()
        if not prompt_file.exists():
            raise typer.BadParameter(f"prompt no existe: {prompt_file}")

    # Engine selection (env -> flag)
    env_pref = engine_from_env()
    eff_engine_name = env_pref or engine_name
    try:
        engine = select_engine(eff_engine_name, claude_bin=claude_bin, codex_bin=codex_bin)
    except (RuntimeError, ValueError) as exc:
        raise typer.BadParameter(str(exc)) from None

    scheduler = (
        BatchScheduler(
            services_per_window=max_services_per_window,
            window_seconds=window_hours * 3600,
            on_event=lambda m: console.print(f"[yellow]{m}[/yellow]"),
        )
        if max_services_per_window > 0
        else None
    )

    schema_path = _ensure_migrate_schema(ws)
    eff_reasoning = _normalize_reasoning_effort(reasoning_effort if engine.name == "codex" else None)

    console.print(
        Panel.fit(
            f"[bold]batch migrate[/bold]\n"
            f"Servicios: {len(services)} · Workers: {workers} · Root: {ws}\n"
            f"Engine: [green]{engine.name}[/green] ({engine.subscription_type}) · "
            f"Model: {model or '(default)'} · Reasoning: {eff_reasoning or '(engine default)'} · "
            f"Checklist: {'NO' if skip_check else 'SI'}\n"
            f"Resume: {'SI' if resume else 'NO'} · Retries: {retries} · "
            f"Window: {max_services_per_window or 'off'}/{window_hours}h",
            border_style="cyan",
        )
    )

    rows: list[BatchRow] = []

    with Progress(
        TextColumn("[bold cyan]{task.description}"),
        BarColumn(),
        TextColumn("{task.completed}/{task.total}"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Migrando", total=len(services))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = [
                pool.submit(
                    _run_service_with_retries,
                    service,
                    lambda svc, attempt: _process_migrate_service(
                        svc,
                        ws,
                        schema_path,
                        engine=engine,
                        model=model,
                        reasoning_effort=eff_reasoning,
                        prompt_file=prompt_file,
                        timeout_minutes=timeout_minutes,
                        run_check=not skip_check,
                        unsafe=unsafe,
                        resume=resume or attempt > 0,
                        scheduler=scheduler,
                    ),
                    retries=retries,
                    workspace_resolver=lambda svc: ws / svc,
                    project_resolver=lambda svc, wsp: _find_project_from_fabrics_metadata(wsp) or _find_migrated_project(wsp, svc),
                    run_kind="migrate",
                )
                for service in services
            ]
            for future in as_completed(futures):
                rows.append(future.result())
                progress.advance(task)

    rows.sort(key=lambda r: r.service)
    _render_table("migrate", rows, MIGRATE_FIELD_ORDER)
    md = _write_markdown_report("migrate", rows, ws, MIGRATE_FIELD_ORDER)
    console.print(f"\n[bold]Reporte:[/bold] [cyan]{md}[/cyan]")


@app.command("engines")
def batch_engines(
    claude_bin: Annotated[
        str, typer.Option("--claude-bin", help="Binario de Claude Code CLI")
    ] = "claude",
    codex_bin: Annotated[
        str, typer.Option("--codex-bin", help="Binario de Codex CLI")
    ] = "codex",
) -> None:
    """Muestra que engines AI headless estan disponibles para `batch migrate`."""
    status = available_engines(claude_bin=claude_bin, codex_bin=codex_bin)
    table = Table(title="Engines disponibles", title_style="bold cyan")
    table.add_column("Engine")
    table.add_column("Subscription")
    table.add_column("Status")
    table.add_column("Detalle")
    names = {"claude": "Claude Max", "codex": "ChatGPT Plus/Pro"}
    for key, (ok, reason) in status.items():
        style = "green" if ok else "red"
        table.add_row(
            key,
            names.get(key, "?"),
            f"[{style}]{'OK' if ok else 'NO'}[/{style}]",
            reason,
        )
    console.print(table)
    available = [k for k, (ok, _) in status.items() if ok]
    if available:
        marker = " (prioridad Claude)" if "claude" in available else ""
        console.print(
            f"\n[bold]Auto-select por defecto:[/bold] "
            f"[green]{available[0]}[/green]{marker}"
        )
    else:
        console.print(
            "\n[red]Ningun engine disponible.[/red] Correr "
            "`capamedia install` + `capamedia auth bootstrap`."
        )
