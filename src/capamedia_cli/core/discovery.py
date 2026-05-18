"""Discovery workbook support for OLA edge-case extraction.

The canonical workbook is packaged with the CLI, while explicit paths, the
environment variable, and workspace-local copies remain supported as overrides.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse
import xml.etree.ElementTree as ET

DISCOVERY_DEFAULT_SHEET = "Validacion de servicios"
DISCOVERY_WORKBOOK_NAME = "Discovery_Servicios_Complejidad OLA 1.xlsx"
DISCOVERY_ENV_VAR = "CAPAMEDIA_DISCOVERY_XLSX"
DISCOVERY_DATA_ROOT = Path(__file__).resolve().parent.parent / "data" / "discovery"
_SERVICE_RE = re.compile(r"\b(?:orq|ws)[a-z]+[0-9]{4}\b", re.IGNORECASE)


@dataclass(frozen=True)
class DiscoveryEdgeCase:
    code: str
    title: str
    evidence: str
    severity: str = "medium"


@dataclass(frozen=True)
class DiscoveryEntry:
    service: str
    migrated_name: str = ""
    tribe: str = ""
    acronym: str = ""
    technology: str = ""
    service_type: str = ""
    complexity: str = ""
    service_weight: str = ""
    integrations: str = ""
    methods: str = ""
    observations: str = ""
    deprecated_notes: str = ""
    link_wsdl: str = ""
    link_code: str = ""
    spec_repo: str = ""
    spec_path: str = ""
    code_repo: str = ""
    weight_flags: list[str] = field(default_factory=list)
    edge_cases: list[DiscoveryEdgeCase] = field(default_factory=list)


@dataclass(frozen=True)
class DiscoverySpecArtifact:
    path: Path
    kind: str


@dataclass(frozen=True)
class DiscoverySpecProbe:
    status: str
    repo_dir: Path | None = None
    artifacts: list[DiscoverySpecArtifact] = field(default_factory=list)
    resolved_path: str = ""
    requested_path: str = ""
    error: str = ""


@dataclass(frozen=True)
class SpecBoundaryCase:
    field: str
    constraint: str
    invalid_value: str
    source: str
    reason: str


@dataclass(frozen=True)
class DiscoveryWorkspaceContext:
    root: Path
    service_name: str = ""
    legacy_path: Path | None = None
    migrated_path: Path | None = None


_EDGE_PATTERNS: list[tuple[str, str, str, str]] = [
    (
        "deprecated_or_repoint",
        "Metodo/servicio deprecado o reapuntamiento",
        r"deprecad|no se consume|no se requiere migrar|reapunt|descarta",
        "high",
    ),
    (
        "mq_or_event",
        "Flujo MQ/eventos",
        r"\bMQ\b|CE_EVENTOS|cola|mensaje",
        "high",
    ),
    (
        "external_provider",
        "Proveedor externo/integrador",
        r"interdin|detectid|banred|integrador|proveedor|externa|bpintegrador",
        "high",
    ),
    (
        "cache_or_config_file",
        "Cache/configurable/XML/properties",
        r"cache|shared row|configuraciones/|atributosproducto\.xml|properties|servicio configurable",
        "medium",
    ),
    (
        "crypto_pdf_library",
        "Cifrado/PDF/libreria especial",
        r"cifrad|encript|tcsprovider|itext|pdf",
        "medium",
    ),
    (
        "same_name_bus_was_or_missing_source",
        "Mismo nombre BUS/WAS o fuente faltante",
        r"mismo nombre|no hay las fuentes|\.war|dmgr",
        "high",
    ),
    (
        "regulatory_or_external_datapower",
        "Regulatorio o DataPower externo",
        r"regulatorio|datapower externo|serviciosexternos",
        "medium",
    ),
    (
        "tx_description_validation",
        "Validacion semantica de TX",
        r"descripciones de las tx",
        "high",
    ),
    (
        "conditional_routing",
        "Ruteo condicional",
        r"boolean|tipoconsulta|valor\s+booleano|solo\s+una\s+ump",
        "medium",
    ),
    (
        "out_of_ola_dependency",
        "Dependencia fuera de OLA",
        r"no estuvo contemplado en la ola\s*1",
        "medium",
    ),
    (
        "missing_source_request",
        "Fuente pendiente/RITM",
        r"\bRITM\d+|solicitar\s+.*\.esql|se cierra el servicio",
        "high",
    ),
]


def _norm(value: object) -> str:
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _service_from_text(text: str) -> str:
    match = _SERVICE_RE.search(text)
    return match.group(0).lower() if match else ""


def service_suffix_key(value: str) -> str:
    """Return the stable service key used for fuzzy spec lookup.

    Azure spec folders can be renamed from one acronym to another, for example
    `mdw-msa-sp-wsreglas0010` -> `csg-msa-sp-wsreglas0010`. The `ws...0000`
    suffix is the durable identifier.
    """
    return _service_from_text(value)


def spec_parent_path(spec_path: str) -> str:
    """Return the parent folder that can be sparse-checked out for lookup."""
    cleaned = spec_path.strip().strip("/")
    if "/" not in cleaned:
        return ""
    return cleaned.rsplit("/", 1)[0]


def rank_spec_candidate(candidate_name: str, service_key: str, acronym: str = "") -> int:
    """Score a spec folder candidate for a service suffix.

    Higher is better. Exact suffix match wins even if the acronym differs.
    """
    normalized_candidate = _norm(candidate_name)
    normalized_service = _norm(service_key)
    if not normalized_service or normalized_service not in normalized_candidate:
        return 0
    score = 80
    if normalized_candidate.endswith(normalized_service):
        score += 20
    if acronym and _norm(acronym) and normalized_candidate.startswith(_norm(acronym)):
        score += 5
    return score


def _find_child_containing_service(parent: Path, service_name: str) -> Path | None:
    if not parent.exists() or not service_name:
        return None
    target = service_name.lower()
    try:
        for child in parent.iterdir():
            if child.is_dir() and target in child.name.lower():
                return child
    except OSError:
        return None
    return None


def detect_discovery_workspace(start: Path) -> DiscoveryWorkspaceContext:
    """Infer service/workspace from a CapaMedia folder with legacy/destino."""
    current = start.resolve()
    service_hint = _service_from_text(current.name) or _service_from_text(str(current))

    for base in [current, *current.parents]:
        legacy_root = base / "legacy"
        destino_root = base / "destino"
        if not legacy_root.exists() and not destino_root.exists():
            continue

        service_name = service_hint
        migrated_path = _find_child_containing_service(destino_root, service_name)
        legacy_path = _find_child_containing_service(legacy_root, service_name)

        if not service_name:
            for root in (destino_root, legacy_root):
                if not root.exists():
                    continue
                try:
                    first = next((p for p in root.iterdir() if p.is_dir()), None)
                except OSError:
                    first = None
                if first:
                    service_name = _service_from_text(first.name)
                    break
            migrated_path = _find_child_containing_service(destino_root, service_name)
            legacy_path = _find_child_containing_service(legacy_root, service_name)

        return DiscoveryWorkspaceContext(
            root=base,
            service_name=service_name,
            legacy_path=legacy_path,
            migrated_path=migrated_path,
        )

    return DiscoveryWorkspaceContext(root=current, service_name=service_hint)


def bundled_discovery_workbook() -> Path | None:
    """Return the packaged canonical Discovery workbook, if present."""
    candidate = DISCOVERY_DATA_ROOT / DISCOVERY_WORKBOOK_NAME
    return candidate if candidate.exists() else None


def find_discovery_workbook(start: Path | None = None, explicit: Path | None = None) -> Path | None:
    """Find the canonical Discovery workbook with local overrides first."""
    if explicit:
        return explicit if explicit.exists() else None

    import os

    env_value = os.environ.get(DISCOVERY_ENV_VAR, "").strip()
    if env_value:
        env_path = Path(env_value).expanduser()
        if env_path.exists():
            return env_path

    base = (start or Path.cwd()).resolve()
    candidates: list[Path] = []
    for parent in [base, *base.parents]:
        candidates.extend(
            [
                parent / DISCOVERY_WORKBOOK_NAME,
                parent / ".capamedia" / DISCOVERY_WORKBOOK_NAME,
                parent / "discovery" / DISCOVERY_WORKBOOK_NAME,
            ]
        )

    bundled = bundled_discovery_workbook()
    if bundled is not None:
        candidates.append(bundled)

    home = Path.home()
    candidates.append(home / "Downloads" / DISCOVERY_WORKBOOK_NAME)

    seen: set[Path] = set()
    for candidate in candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        if candidate.exists():
            return candidate
    return None


def _header_index(headers: list[object], contains: str) -> int | None:
    needle = _norm(contains)
    for idx, header in enumerate(headers):
        if needle in _norm(header):
            return idx
    return None


def _cell_link(cell) -> str:
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()
    return _text(cell.value)


def parse_azure_repo_name(url: str) -> str:
    match = re.search(r"/_git/([^/?#]+)", url)
    return unquote(match.group(1)) if match else ""


def parse_azure_path(url: str) -> str:
    parsed = urlparse(url)
    values = parse_qs(parsed.query).get("path", [])
    return unquote(values[0]) if values else ""


def _derive_migrated_name(row_values: list[object], headers: list[object]) -> str:
    nuevo_idx = _header_index(headers, "nuevo nombre")
    value = _text(row_values[nuevo_idx]) if nuevo_idx is not None else ""
    if value and not value.startswith("="):
        return value

    acr_idx = _header_index(headers, "acronimo")
    svc_idx = _header_index(headers, "servicio")
    acronym = _text(row_values[acr_idx]).lower() if acr_idx is not None else ""
    service = _text(row_values[svc_idx]).lower() if svc_idx is not None else ""
    if acronym and service:
        return f"{acronym}-msa-sp-{service}"
    return value


def _collect_weight_flags(row_values: list[object], headers: list[object]) -> list[str]:
    flags: list[str] = []
    for idx, header in enumerate(headers):
        if _norm(header) != "peso":
            continue
        value = _text(row_values[idx])
        if value not in {"Alta", "Media"}:
            continue
        previous = _text(headers[idx - 1]) if idx > 0 else "criterio"
        flags.append(f"{previous}: {value}")
    return flags


def _first_evidence(pattern: str, fields: list[tuple[str, str]]) -> str:
    rx = re.compile(pattern, re.IGNORECASE)
    for label, value in fields:
        match = rx.search(value)
        if not match:
            continue
        line = next((ln.strip() for ln in value.splitlines() if rx.search(ln)), value)
        return f"{label}: {line[:220]}"
    return ""


def classify_edge_cases(
    *,
    integrations: str = "",
    methods: str = "",
    observations: str = "",
    deprecated_notes: str = "",
    cache: str = "",
    cache_source: str = "",
    providers: str = "",
) -> list[DiscoveryEdgeCase]:
    fields = [
        ("Integraciones", integrations),
        ("Metodos", methods),
        ("Observacion", observations),
        ("Tecnologia deprecada", deprecated_notes),
        ("Cache", cache),
        ("Fuente cache", cache_source),
        ("Proveedores", providers),
    ]
    cases: list[DiscoveryEdgeCase] = []
    for code, title, pattern, severity in _EDGE_PATTERNS:
        evidence = _first_evidence(pattern, fields)
        if evidence:
            cases.append(DiscoveryEdgeCase(code=code, title=title, evidence=evidence, severity=severity))
    return cases


_XSD_FACET_ORDER = (
    "length",
    "minLength",
    "maxLength",
    "pattern",
    "enumeration",
    "totalDigits",
    "fractionDigits",
    "minInclusive",
    "maxInclusive",
)


def _xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _clean_qname(value: str) -> str:
    return value.split(":", 1)[-1] if value else ""


def _parse_xml(path: Path) -> ET.Element | None:
    try:
        return ET.fromstring(path.read_text(encoding="utf-8", errors="replace"))
    except (OSError, ET.ParseError):
        return None


def _restriction_facets(simple_type: ET.Element) -> dict[str, list[str]]:
    facets: dict[str, list[str]] = {}
    for node in simple_type.iter():
        if _xml_local_name(node.tag) != "restriction":
            continue
        for child in list(node):
            facet = _xml_local_name(child.tag)
            if facet not in _XSD_FACET_ORDER:
                continue
            value = child.attrib.get("value", "").strip()
            if value:
                facets.setdefault(facet, []).append(value)
    return facets


def _constraint_text(facets: dict[str, list[str]]) -> str:
    parts: list[str] = []
    for facet in _XSD_FACET_ORDER:
        values = facets.get(facet)
        if not values:
            continue
        parts.append(f"{facet}={'|'.join(values)}")
    return "; ".join(parts)


def _int_value(value: str) -> int | None:
    try:
        return int(value)
    except ValueError:
        return None


def _number_value(value: str) -> float | None:
    try:
        return float(value)
    except ValueError:
        return None


def _invalid_value_for_facets(facets: dict[str, list[str]]) -> tuple[str, str]:
    if facets.get("enumeration"):
        return "__INVALID_ENUM__", "valor fuera de la enumeracion permitida"
    if facets.get("pattern"):
        return "__INVALID_PATTERN__", "valor que no cumple el pattern XSD"
    if facets.get("length"):
        size = _int_value(facets["length"][0])
        if size is not None:
            return "X" * min(size + 1, 80), f"longitud distinta de {size}"
    if facets.get("maxLength"):
        size = _int_value(facets["maxLength"][0])
        if size is not None:
            return "X" * min(size + 1, 80), f"longitud mayor a {size}"
    if facets.get("minLength"):
        size = _int_value(facets["minLength"][0])
        if size is not None:
            return "" if size > 0 else "X", f"longitud menor a {size}"
    if facets.get("totalDigits"):
        digits = _int_value(facets["totalDigits"][0])
        if digits is not None:
            return "9" * min(digits + 1, 80), f"mas de {digits} digitos"
    if facets.get("fractionDigits"):
        digits = _int_value(facets["fractionDigits"][0])
        if digits is not None:
            return "0." + ("1" * min(digits + 1, 20)), f"mas de {digits} decimales"
    if facets.get("minInclusive"):
        value = _number_value(facets["minInclusive"][0])
        if value is not None:
            return str(int(value - 1) if value.is_integer() else value - 1), "valor menor al minimo"
    if facets.get("maxInclusive"):
        value = _number_value(facets["maxInclusive"][0])
        if value is not None:
            return str(int(value + 1) if value.is_integer() else value + 1), "valor mayor al maximo"
    return "INVALID", "valor invalido segun restriccion XSD"


def extract_spec_boundary_cases(artifacts: list[DiscoverySpecArtifact]) -> list[SpecBoundaryCase]:
    """Extract QA boundary/overflow cases from WSDL/XSD simpleType facets.

    The spec repo is the authoritative source for XML contract limits. We read
    both WSDL inline schemas and external XSDs, then generate one invalid value
    per constrained field so QA prompts can exercise overflow/negative cases.
    """
    roots: list[tuple[Path, ET.Element]] = []
    for artifact in artifacts:
        if artifact.path.suffix.lower() not in {".wsdl", ".xsd"}:
            continue
        root = _parse_xml(artifact.path)
        if root is not None:
            roots.append((artifact.path, root))

    named_types: dict[str, dict[str, list[str]]] = {}
    for _, root in roots:
        for node in root.iter():
            if _xml_local_name(node.tag) != "simpleType":
                continue
            name = node.attrib.get("name", "").strip()
            if name:
                facets = _restriction_facets(node)
                if facets:
                    named_types[name] = facets

    cases: list[SpecBoundaryCase] = []
    seen: set[tuple[str, str]] = set()
    for path, root in roots:
        for element in root.iter():
            if _xml_local_name(element.tag) != "element":
                continue
            field = element.attrib.get("name") or _clean_qname(element.attrib.get("ref", ""))
            if not field:
                continue

            facets: dict[str, list[str]] = {}
            type_name = _clean_qname(element.attrib.get("type", ""))
            if type_name in named_types:
                facets = named_types[type_name]
            else:
                for child in list(element):
                    if _xml_local_name(child.tag) == "simpleType":
                        facets = _restriction_facets(child)
                        break
            if not facets:
                continue

            constraint = _constraint_text(facets)
            key = (field, constraint)
            if key in seen:
                continue
            seen.add(key)
            invalid_value, reason = _invalid_value_for_facets(facets)
            cases.append(
                SpecBoundaryCase(
                    field=field,
                    constraint=constraint,
                    invalid_value=invalid_value,
                    source=path.name,
                    reason=reason,
                )
            )
    return cases


def load_discovery_entry(
    workbook_path: Path,
    service_name: str,
    *,
    sheet_name: str = DISCOVERY_DEFAULT_SHEET,
) -> DiscoveryEntry | None:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency declared in pyproject
        raise RuntimeError("openpyxl no disponible para leer discovery .xlsx") from exc

    wb = load_workbook(workbook_path, data_only=True, read_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Hoja '{sheet_name}' no existe en {workbook_path}")
    ws = wb[sheet_name]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    required = {
        "servicio": _header_index(headers, "servicio"),
        "nuevo": _header_index(headers, "nuevo nombre"),
        "tribu": _header_index(headers, "tribu"),
        "acronimo": _header_index(headers, "acronimo"),
        "tecnologia": _header_index(headers, "tecnologia"),
        "tipo": _header_index(headers, "tipo"),
        "integraciones": _header_index(headers, "integraciones consume"),
        "cache": _header_index(headers, "cache adicional"),
        "cache_source": _header_index(headers, "archivo o servicio"),
        "providers": _header_index(headers, "proveedores externos"),
        "methods": _header_index(headers, "metodos que expone"),
        "service_weight": _header_index(headers, "peso del servicio"),
        "complexity": _header_index(headers, "complejidad del servicio"),
        "observations": _header_index(headers, "observacion"),
        "link_wsdl": _header_index(headers, "link wsdl"),
        "link_code": _header_index(headers, "link codigo"),
        "deprecated": _header_index(headers, "consumen tecnologia deprecada"),
    }
    if required["servicio"] is None:
        raise ValueError("No se encontro columna Servicio en discovery")

    target = _norm(service_name)
    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        service = _text(values[required["servicio"]])
        if not service:
            continue
        migrated_name = _derive_migrated_name(values, headers)
        if target not in {_norm(service), _norm(migrated_name)}:
            continue

        def value(name: str, row_values: list[object] = values) -> str:
            idx = required[name]
            return _text(row_values[idx]) if idx is not None else ""

        link_wsdl = _cell_link(row[required["link_wsdl"]]) if required["link_wsdl"] is not None else ""
        link_code = _cell_link(row[required["link_code"]]) if required["link_code"] is not None else ""
        integrations = value("integraciones")
        methods = value("methods")
        observations = value("observations")
        deprecated_notes = value("deprecated")
        cache = value("cache")
        cache_source = value("cache_source")
        providers = value("providers")

        return DiscoveryEntry(
            service=service,
            migrated_name=migrated_name,
            tribe=value("tribu"),
            acronym=value("acronimo"),
            technology=value("tecnologia"),
            service_type=value("tipo"),
            complexity=value("complexity"),
            service_weight=value("service_weight"),
            integrations=integrations,
            methods=methods,
            observations=observations,
            deprecated_notes=deprecated_notes,
            link_wsdl=link_wsdl,
            link_code=link_code,
            spec_repo=parse_azure_repo_name(link_wsdl),
            spec_path=parse_azure_path(link_wsdl),
            code_repo=parse_azure_repo_name(link_code),
            weight_flags=_collect_weight_flags(values, headers),
            edge_cases=classify_edge_cases(
                integrations=integrations,
                methods=methods,
                observations=observations,
                deprecated_notes=deprecated_notes,
                cache=cache,
                cache_source=cache_source,
                providers=providers,
            ),
        )
    return None


def render_discovery_markdown(
    entry: DiscoveryEntry,
    *,
    spec_probe: DiscoverySpecProbe | None = None,
    copied_artifacts: list[Path] | None = None,
) -> str:
    lines: list[str] = []
    spec_boundary_cases = (
        extract_spec_boundary_cases(spec_probe.artifacts)
        if spec_probe and spec_probe.artifacts
        else []
    )
    lines.append("## Discovery / edge cases")
    lines.append("")
    lines.append(f"- **Servicio discovery:** `{entry.service}`")
    if entry.migrated_name:
        lines.append(f"- **Nombre migrado discovery:** `{entry.migrated_name}`")
    lines.append(f"- **Tipo / Tecnologia:** `{entry.service_type or '?'} / {entry.technology or '?'}`")
    lines.append(f"- **Complejidad discovery:** `{entry.complexity or '?'}` (peso `{entry.service_weight or '?'}`)")
    lines.append(f"- **Spec repo:** `{entry.spec_repo or '?'}`")
    resolved_spec_path = spec_probe.resolved_path if spec_probe and spec_probe.resolved_path else entry.spec_path
    lines.append(f"- **Spec path:** `{resolved_spec_path or '?'}`")
    if spec_probe and spec_probe.requested_path and spec_probe.resolved_path and spec_probe.requested_path != spec_probe.resolved_path:
        lines.append(f"- **Spec path solicitado:** `{spec_probe.requested_path}`")
        lines.append(f"- **Spec path resuelto:** `{spec_probe.resolved_path}`")
    lines.append(f"- **Code repo:** `{entry.code_repo or '?'}`")
    lines.append("")

    if entry.methods:
        lines.append("### Metodos discovery")
        lines.append("")
        for method in entry.methods.splitlines():
            if method.strip():
                lines.append(f"- {method.strip()}")
        lines.append("")

    if entry.weight_flags:
        lines.append("### Flags de complejidad")
        lines.append("")
        for flag in entry.weight_flags:
            lines.append(f"- {flag}")
        lines.append("")

    lines.append("### DISCOVERY_EDGE_CASES")
    lines.append("")
    lines.append("```text")
    lines.append("DISCOVERY_EDGE_CASES:")
    lines.append(f"- spec_path: {resolved_spec_path or '<not_provided>'}")
    lines.append(f"- code_repo: {entry.code_repo or '<not_provided>'}")
    if spec_probe and spec_probe.artifacts:
        artifact_names = ", ".join(artifact.path.name for artifact in spec_probe.artifacts)
        lines.append(f"- spec_artifacts: {artifact_names}")
    if spec_boundary_cases:
        boundary_summary = ", ".join(
            f"{case.field}({case.constraint})" for case in spec_boundary_cases[:20]
        )
        lines.append(f"- spec_boundary_cases: {boundary_summary}")
    if copied_artifacts:
        copied_names = ", ".join(str(path) for path in copied_artifacts)
        lines.append(f"- copied_artifacts: {copied_names}")
    edge_codes = ", ".join(case.code for case in entry.edge_cases) or "<none>"
    lines.append(f"- edge_cases: {edge_codes}")
    lines.append("- test_case_source: not_probed")
    lines.append("```")
    lines.append("")

    lines.append("### Edge cases")
    lines.append("")
    if entry.edge_cases:
        lines.append("| Codigo | Severidad | Evidencia |")
        lines.append("|---|---|---|")
        for case in entry.edge_cases:
            lines.append(f"| `{case.code}` | `{case.severity}` | {case.evidence} |")
    else:
        lines.append("_(ninguno detectado por discovery)_")
    lines.append("")

    if entry.edge_cases:
        lines.append("### Discovery edge-case coverage")
        lines.append("")

    if spec_boundary_cases:
        lines.append("### Casos de desborde desde WSDL/XSD")
        lines.append("")
        lines.append("| Campo | Restriccion XSD | Valor invalido sugerido | Fuente | Motivo |")
        lines.append("|---|---|---|---|---|")
        for case in spec_boundary_cases:
            constraint = case.constraint.replace("|", "\\|")
            invalid_value = case.invalid_value.replace("|", "\\|")
            reason = case.reason.replace("|", "\\|")
            lines.append(
                f"| `{case.field}` | `{constraint}` | `{invalid_value}` | "
                f"`{case.source}` | {reason} |"
            )
        lines.append("")

    if spec_probe:
        lines.append("### Artefactos WSDL/XSD")
        lines.append("")
        lines.append(f"- Estado probe: `{spec_probe.status}`")
        if spec_probe.requested_path:
            lines.append(f"- Path solicitado: `{spec_probe.requested_path}`")
        if spec_probe.resolved_path:
            lines.append(f"- Path resuelto: `{spec_probe.resolved_path}`")
        if spec_probe.error:
            lines.append(f"- Error: `{spec_probe.error}`")
        if spec_probe.artifacts:
            lines.append("")
            lines.append("| Archivo | Tipo | Origen |")
            lines.append("|---|---|---|")
            for artifact in spec_probe.artifacts:
                lines.append(f"| `{artifact.path.name}` | `{artifact.kind}` | `{artifact.path}` |")
        else:
            lines.append("")
            lines.append("No se materializaron artefactos WSDL/XSD.")
        if copied_artifacts:
            lines.append("")
            lines.append("#### Copiados al servicio migrado")
            lines.append("")
            for path in copied_artifacts:
                lines.append(f"- `{path}`")
        lines.append("")
        lines.append("| Codigo | Decision | Implementacion / test |")
        lines.append("|---|---|---|")
        for case in entry.edge_cases:
            lines.append(f"| `{case.code}` | PENDIENTE | <pendiente_validar> |")
        lines.append("")

    if entry.observations:
        lines.append("### Observacion Discovery")
        lines.append("")
        lines.append(entry.observations)
        lines.append("")

    if entry.integrations:
        lines.append("### Integraciones / consume")
        lines.append("")
        lines.append("```text")
        lines.append(entry.integrations[:3000])
        lines.append("```")
        lines.append("")

    if entry.link_wsdl:
        lines.append(f"- LINK WSDL: {entry.link_wsdl}")
    if entry.link_code:
        lines.append(f"- LINK CODIGO: {entry.link_code}")
    lines.append("")
    return "\n".join(lines)
